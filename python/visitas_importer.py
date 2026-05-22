import argparse
import os
import re
import sys
import time
from datetime import date, datetime, timedelta

import mysql.connector
from dotenv import load_dotenv
from openpyxl import load_workbook


def log_line(msg: str):
    p = os.getenv("VISITAS_LOG_PATH", "").strip()
    if not p:
        return
    try:
        with open(p, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.utcnow().isoformat(timespec='seconds')}Z] {msg}\n")
    except Exception:
        pass


def connect_db():
    host = os.getenv("DB_HOST", "127.0.0.1")
    port = int(os.getenv("DB_PORT", "3306"))
    user = os.getenv("DB_USER", "")
    password = os.getenv("DB_PASSWORD", "")
    database = os.getenv("DB_NAME", "")
    return mysql.connector.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        autocommit=False,
    )


def to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_int(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if v != v:
            return None
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d\-]+", "", s)
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def to_decimal_text(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return None
        return str(v)
    s = str(v).strip()
    return s if s else None


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return None
        base = date(1899, 12, 30)
        try:
            return base + timedelta(days=int(v))
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def first_day_month(d: date):
    return date(d.year, d.month, 1)


def get_cell(row_vals, idx):
    if idx is None:
        return None
    if idx < 0:
        return None
    if idx >= len(row_vals):
        return None
    return row_vals[idx]


def job_update(cur, job_id: str, **fields):
    if not fields:
        return
    set_parts = []
    vals = []
    for k, v in fields.items():
        set_parts.append(f"{k} = %s")
        vals.append(v)
    vals.append(job_id)
    cur.execute(f"UPDATE visitas_import_jobs SET {', '.join(set_parts)} WHERE id = %s", vals)


def load_config(cur, config_id: int):
    cur.execute("SELECT * FROM visitas_import_configs WHERE id = %s LIMIT 1", [config_id])
    r = cur.fetchone()
    if not r:
        raise Exception(
            f"No existe configuración de columnas (visitas_import_configs) con id={config_id}."
        )
    cols = [d[0] for d in cur.description]
    return dict(zip(cols, r))


RAW_INSERT_SQL = """
INSERT INTO visitas_raw (
  job_id, ubigeo, etapa_mes, dni_nino, etapa_text, visitas_completas_edad,
  fecha_intervencion, dispositivo, estado_intervencion, latitud, longitud
) VALUES (
  %s,%s,%s,%s,%s,%s,
  %s,%s,%s,%s,%s
)
"""


MONTH_UPSERT_SQL = """
INSERT INTO visitas_mensual (
  ubigeo, etapa_mes, dni_nino, expected_visits, visitas_count, fecha_v1, fecha_v2, fecha_v3,
  completa, oportuna, cumple, georef_visits, has_georef, flag_no_encontrado, flag_rechazado
) VALUES (
  %s,%s,%s,%s,%s,%s,%s,%s,
  %s,%s,%s,%s,%s,%s,%s
)
ON DUPLICATE KEY UPDATE
  expected_visits = VALUES(expected_visits),
  visitas_count = VALUES(visitas_count),
  fecha_v1 = VALUES(fecha_v1),
  fecha_v2 = VALUES(fecha_v2),
  fecha_v3 = VALUES(fecha_v3),
  completa = VALUES(completa),
  oportuna = VALUES(oportuna),
  cumple = VALUES(cumple),
  georef_visits = VALUES(georef_visits),
  has_georef = VALUES(has_georef),
  flag_no_encontrado = VALUES(flag_no_encontrado),
  flag_rechazado = VALUES(flag_rechazado)
"""


def calc_month_summary(events):
    visit_dates = []
    expected = None
    georef_visits = 0
    flag_no = 0
    flag_rech = 0

    for ev in events:
        etapa_text = (ev.get("etapa_text") or "").strip().lower()
        estado = (ev.get("estado_intervencion") or "").strip().lower()

        if "no encontrado" in etapa_text or "no encontrado" in estado:
            flag_no = 1
        if "rechaz" in etapa_text or "rechaz" in estado:
            flag_rech = 1

        if etapa_text and not etapa_text.startswith("visita"):
            continue

        d = ev.get("fecha_intervencion")
        if isinstance(d, date):
            visit_dates.append(d)

        e = ev.get("visitas_completas_edad")
        if isinstance(e, int):
            expected = e if expected is None else max(expected, e)

        lat = ev.get("latitud")
        lon = ev.get("longitud")
        if lat is not None and lon is not None:
            georef_visits += 1

    visit_dates = sorted(set(visit_dates))
    visitas_count = len(visit_dates)
    expected_visits = expected if isinstance(expected, int) and expected >= 0 else None

    completa = 0
    if expected_visits is None:
        completa = 0
    else:
        completa = 1 if visitas_count >= expected_visits and expected_visits > 0 else 0

    oportuna = 0
    if completa == 1:
        if visitas_count <= 1:
            oportuna = 1
        else:
            ok = True
            for i in range(1, len(visit_dates)):
                diff = (visit_dates[i] - visit_dates[i - 1]).days
                if diff < 7 or diff > 10:
                    ok = False
                    break
            oportuna = 1 if ok else 0
    cumple = 1 if (completa == 1 and oportuna == 1 and flag_no == 0 and flag_rech == 0) else 0

    v1 = visit_dates[0] if len(visit_dates) >= 1 else None
    v2 = visit_dates[1] if len(visit_dates) >= 2 else None
    v3 = visit_dates[2] if len(visit_dates) >= 3 else None

    has_georef = 1 if georef_visits > 0 else 0
    return {
        "expected_visits": expected_visits,
        "visitas_count": visitas_count,
        "fecha_v1": v1,
        "fecha_v2": v2,
        "fecha_v3": v3,
        "completa": completa,
        "oportuna": oportuna,
        "cumple": cumple,
        "georef_visits": georef_visits,
        "has_georef": has_georef,
        "flag_no_encontrado": flag_no,
        "flag_rechazado": flag_rech,
    }


def run_import(job_id: str, file_path: str, config_id: int):
    load_dotenv(os.getenv("VISITAS_DOTENV", ".env.local"), override=False)
    load_dotenv(override=False)

    log_line(f"Job {job_id} starting. File: {file_path} config_id={config_id}")
    db = connect_db()
    cur = db.cursor()

    cfg = load_config(cur, config_id)
    db.commit()

    job_update(
        cur,
        job_id,
        status="running",
        progress=0,
        started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        message="Leyendo Excel...",
    )
    db.commit()

    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet_index = int(cfg.get("sheet_index") or 0)
    if sheet_index < 0 or sheet_index >= len(wb.worksheets):
        sheet_index = 0
    ws = wb.worksheets[sheet_index]

    start_row = int(cfg.get("start_row") or 1)
    if start_row < 1:
        start_row = 1

    events = []
    for row_vals in ws.iter_rows(min_row=start_row, values_only=True):
        dni = to_text(get_cell(row_vals, cfg.get("col_dni_nino")))
        if not dni:
            continue

        ubigeo = to_int(get_cell(row_vals, cfg.get("col_ubigeo")))
        if ubigeo is None:
            continue

        fecha = to_date(get_cell(row_vals, cfg.get("col_fecha_intervencion")))
        if fecha is None:
            continue
        etapa_mes = first_day_month(fecha)

        etapa_text = to_text(get_cell(row_vals, cfg.get("col_etapa_text")))
        visitas_completas = to_int(get_cell(row_vals, cfg.get("col_visitas_completas")))
        dispositivo = to_text(get_cell(row_vals, cfg.get("col_dispositivo")))
        estado_int = to_text(get_cell(row_vals, cfg.get("col_estado_intervencion")))
        lat = to_decimal_text(get_cell(row_vals, cfg.get("col_latitud")))
        lon = to_decimal_text(get_cell(row_vals, cfg.get("col_longitud")))

        events.append(
            {
                "ubigeo": ubigeo,
                "dni_nino": dni,
                "etapa_mes": etapa_mes,
                "etapa_text": etapa_text,
                "visitas_completas_edad": visitas_completas,
                "fecha_intervencion": fecha,
                "dispositivo": dispositivo,
                "estado_intervencion": estado_int,
                "latitud": lat,
                "longitud": lon,
            }
        )

    total_rows = len(events)
    if total_rows == 0:
        job_update(
            cur,
            job_id,
            status="failed",
            finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            message="No se encontraron registros (DNI/Ubigeo/Fecha vacíos). Revisa columnas y fila inicio.",
        )
        db.commit()
        return

    uniq_pairs = sorted(set([(e["ubigeo"], e["etapa_mes"]) for e in events]))
    log_line(f"Found rows={total_rows} month_pairs={len(uniq_pairs)}")

    job_update(cur, job_id, total_rows=total_rows, processed_rows=0, inserted_rows=0, progress=0)
    db.commit()

    cur.execute("SET SESSION sql_mode = ''")

    placeholders = ",".join(["(%s,%s)"] * len(uniq_pairs))
    del_vals = []
    for ub, em in uniq_pairs:
        del_vals.append(ub)
        del_vals.append(em)

    if uniq_pairs:
        cur.execute(f"DELETE FROM visitas_mensual WHERE (ubigeo, etapa_mes) IN ({placeholders})", del_vals)
        cur.execute(f"DELETE FROM visitas_raw WHERE (ubigeo, etapa_mes) IN ({placeholders})", del_vals)
        db.commit()

    batch_size = 1000
    inserted_raw = 0
    processed = 0
    last_tick = time.time()

    for i in range(0, total_rows, batch_size):
        batch = events[i : i + batch_size]
        vals = []
        for e in batch:
            vals.append(
                (
                    job_id,
                    e["ubigeo"],
                    e["etapa_mes"],
                    e["dni_nino"],
                    e["etapa_text"],
                    e["visitas_completas_edad"],
                    e["fecha_intervencion"],
                    e["dispositivo"],
                    e["estado_intervencion"],
                    e["latitud"],
                    e["longitud"],
                )
            )
        cur.executemany(RAW_INSERT_SQL, vals)
        db.commit()
        inserted_raw += cur.rowcount if cur.rowcount > 0 else 0
        processed += len(batch)

        now = time.time()
        if now - last_tick >= 0.6 or processed >= total_rows:
            progress = int((processed / max(1, total_rows)) * 70)
            job_update(
                cur,
                job_id,
                progress=progress,
                processed_rows=processed,
                inserted_rows=inserted_raw,
                message=f"Insertando visitas... {processed}/{total_rows}",
            )
            db.commit()
            last_tick = now

    job_update(cur, job_id, message="Generando resumen mensual...")
    db.commit()

    grouped = {}
    for e in events:
        k = (e["ubigeo"], e["etapa_mes"], e["dni_nino"])
        grouped.setdefault(k, []).append(e)

    keys = list(grouped.keys())
    total_groups = len(keys)
    inserted_month = 0
    processed_groups = 0
    last_tick = time.time()

    for i in range(0, total_groups, batch_size):
        part = keys[i : i + batch_size]
        vals = []
        for k in part:
            ub, em, dni = k
            summary = calc_month_summary(grouped[k])
            vals.append(
                (
                    ub,
                    em,
                    dni,
                    summary["expected_visits"],
                    summary["visitas_count"],
                    summary["fecha_v1"],
                    summary["fecha_v2"],
                    summary["fecha_v3"],
                    summary["completa"],
                    summary["oportuna"],
                    summary["cumple"],
                    summary["georef_visits"],
                    summary["has_georef"],
                    summary["flag_no_encontrado"],
                    summary["flag_rechazado"],
                )
            )
        cur.executemany(MONTH_UPSERT_SQL, vals)
        db.commit()
        inserted_month += cur.rowcount if cur.rowcount > 0 else 0
        processed_groups += len(part)

        now = time.time()
        if now - last_tick >= 0.6 or processed_groups >= total_groups:
            progress = 70 + int((processed_groups / max(1, total_groups)) * 30)
            job_update(
                cur,
                job_id,
                progress=progress,
                message=f"Resumiendo... {processed_groups}/{total_groups}",
            )
            db.commit()
            last_tick = now

    job_update(
        cur,
        job_id,
        status="done",
        progress=100,
        processed_rows=total_rows,
        inserted_rows=inserted_raw,
        finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        message=f"Completado. Visitas: {inserted_raw} · Resúmenes: {total_groups}",
    )
    db.commit()
    log_line(f"Done. raw={inserted_raw} months={total_groups}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--job", required=True)
    ap.add_argument("--file", required=True)
    ap.add_argument("--config", required=False, type=int, default=1)
    args = ap.parse_args()

    try:
        run_import(args.job, args.file, int(args.config or 1))
    except Exception as e:
        try:
            load_dotenv(os.getenv("VISITAS_DOTENV", ".env.local"), override=False)
            load_dotenv(override=False)
            db = connect_db()
            cur = db.cursor()
            job_update(
                cur,
                args.job,
                status="failed",
                progress=0,
                finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                message=str(e),
            )
            db.commit()
        except Exception:
            pass
        log_line(f"FAILED: {e}")
        raise


if __name__ == "__main__":
    main()

