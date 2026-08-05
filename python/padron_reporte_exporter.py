import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from typing import Any

import mysql.connector
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def log_line(msg: str):
    p = os.getenv("PADRON_REPORTE_LOG_PATH", "").strip()
    if not p:
        return
    try:
        with open(p, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.utcnow().isoformat(timespec='seconds')}Z] {msg}\n")
    except Exception:
        pass


def connect_db():
    host = os.getenv("DB_HOST", "127.0.0.1")
    port = int(os.getenv("DB_PORT", "3306"))
    user = os.getenv("DB_USER", "")
    password = os.getenv("DB_PASSWORD", "")
    database = os.getenv("DB_NAME", "")
    return mysql.connector.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        autocommit=False,
    )


def normalize_etapas(raw):
    etapas = []
    for x in raw or []:
        s = str(x or "").strip()
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            etapas.append(s)
    return list(dict.fromkeys(etapas))


def normalize_ubigeos(raw):
    ub = []
    for x in raw or []:
        try:
            n = int(str(x).strip())
            ub.append(n)
        except Exception:
            pass
    return list(dict.fromkeys(ub))


def job_update(cur, job_id: int, status: str, progress: int, message: str = None, file_path: str = None):
    cur.execute(
        "UPDATE report_export_jobs SET status=%s, progress=%s, message=%s, file_path=%s WHERE id=%s",
        (status, int(progress), message, file_path, int(job_id)),
    )


def build_query(tipo: str, etapas: list, ubigeos: list):
    where = ["TRIM(COALESCE(pn0.tipovd,'')) = %s", f"pn0.etapa IN ({','.join(['%s'] * len(etapas))})"]
    values = [tipo] + etapas[:]

    if len(ubigeos) == 1:
        where.insert(0, "pn0.ubigeo = %s")
        values.insert(0, ubigeos[0])
    elif len(ubigeos) > 1:
        where.insert(0, f"pn0.ubigeo IN ({','.join(['%s'] * len(ubigeos))})")
        values = ubigeos[:] + values

    vr_ubigeo_sql = ""
    if len(ubigeos) == 1:
        vr_ubigeo_sql = "AND vr0.ubigeo = %s"
    elif len(ubigeos) > 1:
        vr_ubigeo_sql = f"AND vr0.ubigeo IN ({','.join(['%s'] * len(ubigeos))})"

    where_pn2 = [w.replace("pn0.", "pn2.") for w in where]
    values_pn2 = values[:]

    query_values = []
    query_values.extend(values)
    query_values.extend(etapas)
    if len(ubigeos) > 0:
        query_values.extend(ubigeos)
    query_values.extend(values_pn2)

    limit = 2000000

    sql = f"""
WITH pdj AS (
  SELECT ubigeo, periodo, job_id
  FROM (
    SELECT
      ubigeo,
      periodo,
      id AS job_id,
      ROW_NUMBER() OVER (
        PARTITION BY ubigeo, periodo
        ORDER BY fecha_corte DESC, created_at DESC, id DESC
      ) AS rn
    FROM padron_dni_import_jobs
    WHERE status = 'done'
  ) x
  WHERE rn = 1
),
pdrx AS (
  SELECT *
  FROM (
    SELECT
      r0.id,
      r0.job_id,
      NULLIF(TRIM(COALESCE(r0.dni,'')), '') AS dni_stored,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[3]'))), '') AS cnv_key,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[5]'))), '') AS dni_key,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[2]'))), '') AS codpad_key,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[8]'))), '') AS nino_appat,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[9]'))), '') AS nino_apmat,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[10]'))), '') AS nino_nombres,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[44]'))), '') AS madre_dni,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[45]'))), '') AS madre_appat,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[46]'))), '') AS madre_apmat,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[47]'))), '') AS madre_nombres,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[48]'))), '') AS madre_celular,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[54]'))), '') AS jefe_dni,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[55]'))), '') AS jefe_appat,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[56]'))), '') AS jefe_apmat,
      NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[57]'))), '') AS jefe_nombres,
      ROW_NUMBER() OVER (
        PARTITION BY
          r0.job_id,
          COALESCE(
            NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[3]'))), ''),
            NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[5]'))), ''),
            NULLIF(TRIM(JSON_UNQUOTE(JSON_EXTRACT(r0.payload, '$[2]'))), ''),
            NULLIF(TRIM(COALESCE(r0.dni,'')), '')
          )
        ORDER BY r0.id DESC
      ) AS rn_key
    FROM padron_dni_raw r0
    JOIN pdj ON pdj.job_id = r0.job_id
    WHERE JSON_VALID(r0.payload)
  ) y
  WHERE rn_key = 1
)
SELECT
  pn.idpn, pn.tipo, pn.rango, pn.ccpp, pn.zona, pn.mz, pn.direccion, pn.referencia, pn.codeess, pn.tipodoc,
  pn.dni,
  COALESCE(
    NULLIF(NULLIF(UPPER(TRIM(COALESCE(pn.nombres,''))), 'NULL'), ''),
    NULLIF(TRIM(CONCAT_WS(' ', pdr_cnv.nino_appat, pdr_cnv.nino_apmat, pdr_cnv.nino_nombres)), ''),
    NULLIF(TRIM(CONCAT_WS(' ', pdr_dni.nino_appat, pdr_dni.nino_apmat, pdr_dni.nino_nombres)), ''),
    NULLIF(TRIM(CONCAT_WS(' ', pdr_cod.nino_appat, pdr_cod.nino_apmat, pdr_cod.nino_nombres)), '')
  ) AS nombres,
  pn.fecha_nac,
  COALESCE(NULLIF(NULLIF(UPPER(TRIM(COALESCE(pn.dnimadre,''))), 'NULL'), ''), pdr_cnv.madre_dni, pdr_dni.madre_dni, pdr_cod.madre_dni) AS dnimadre,
  COALESCE(NULLIF(NULLIF(UPPER(TRIM(COALESCE(pn.appatmadre,''))), 'NULL'), ''), pdr_cnv.madre_appat, pdr_dni.madre_appat, pdr_cod.madre_appat) AS appatmadre,
  COALESCE(NULLIF(NULLIF(UPPER(TRIM(COALESCE(pn.apmatmadre,''))), 'NULL'), ''), pdr_cnv.madre_apmat, pdr_dni.madre_apmat, pdr_cod.madre_apmat) AS apmatmadre,
  COALESCE(NULLIF(NULLIF(UPPER(TRIM(COALESCE(pn.nombresmadre,''))), 'NULL'), ''), pdr_cnv.madre_nombres, pdr_dni.madre_nombres, pdr_cod.madre_nombres) AS nombresmadre,
  COALESCE(NULLIF(NULLIF(UPPER(TRIM(COALESCE(pn.dni_padre,''))), 'NULL'), ''), pdr_cnv.jefe_dni, pdr_dni.jefe_dni, pdr_cod.jefe_dni) AS dni_padre,
  COALESCE(
    NULLIF(NULLIF(UPPER(TRIM(COALESCE(pn.nombre_padre,''))), 'NULL'), ''),
    NULLIF(TRIM(CONCAT_WS(' ', pdr_cnv.jefe_appat, pdr_cnv.jefe_apmat, pdr_cnv.jefe_nombres)), ''),
    NULLIF(TRIM(CONCAT_WS(' ', pdr_dni.jefe_appat, pdr_dni.jefe_apmat, pdr_dni.jefe_nombres)), ''),
    NULLIF(TRIM(CONCAT_WS(' ', pdr_cod.jefe_appat, pdr_cod.jefe_apmat, pdr_cod.jefe_nombres)), '')
  ) AS nombre_padre,
  pn.idocurrencia, pn.idocurrencia2,
  pn.nuevadireccion, pn.nuevareferencia,
  pn.observacion, pn.obspadron,
  pn.actorsocial, pn.responsable,
  pn.telefono,
  COALESCE(NULLIF(NULLIF(UPPER(TRIM(COALESCE(pn.telefonopn,''))), 'NULL'), ''), pdr_cnv.madre_celular, pdr_dni.madre_celular, pdr_cod.madre_celular) AS telefonopn,
  pn.adulto, pn.cantidada,
  pn.etapa, pn.estadovd, pn.fechacita, pn.nrovd,
  pn.eess_ua, pn.departamento, pn.provincia, pn.distrito,
  pn.codqr, pn.modovd, pn.tipovd,
  pn.lat, pn.lon, pn.lat2, pn.long2, pn.lat3, pn.long3,
  pn.fechamodificacion, pn.fechamodificacion2,
  pn.tamisaje, pn.hb, pn.anemia, pn.hierro, pn.tsf, pn.rsf,
  pn.visitadops, pn.sesiondem, pn.tieneps,
  pn.observacion2, pn.fechatamisaje,
  pn.tiposeguro, pn.tipodocum, pn.usuario,
  pn.ubigeo,
  pn.fecha_inicio_vd, pn.fecha_fin_vd,
  vr.estadosvd, vr.estadosvd2, vr.estadosvd3,
  vr.primera_vd, vr.segunda_vd, vr.tercera_vd,
  tam.fecha_atencion AS tam_fecha_atencion,
  tam.hemoglobina AS tam_hemoglobina,
  tam.peso AS tam_peso,
  tam.talla AS tam_talla,
  tam.cie_10 AS tam_cie_10,
  tam.resultado AS tam_resultado,
  tam.hemoglobina1, tam.fecha_atencion1,
  tam.hemoglobina2, tam.fecha_atencion2,
  tam.hemoglobina3, tam.fecha_atencion3,
  pn.resultado, pn.avance,
  pn.fotos, pn.programacion1, pn.padronnominal,
  pn.iddistrito, pn.discapacidad, pn.titular_linea, pn.codigov,
  pn.img_carnet, pn.estado_verificacion, pn.estado, pn.estado_verificado,
  pn.celularseapp, pn.tipodispositivo, pn.estadointervencion, pn.asignacion,
  pn.estadoseguro, pn.fecha_act_seguro, pn.nombre_comercial, pn.hbregistro, pn.ccred,
  TRIM(CONCAT(COALESCE(p_actor.nombrecompleto,''),' ',COALESCE(p_actor.apellidos,''))) AS actor_nombre,
  TRIM(CONCAT(COALESCE(p_resp.nombrecompleto,''),' ',COALESCE(p_resp.apellidos,''))) AS responsable_nombre,
  o1.descripcion AS ocurrencia_desc,
  o2.descripcion AS ocurrencia2_desc
FROM (
  SELECT
    pn0.*,
    ROW_NUMBER() OVER (
      PARTITION BY pn0.ubigeo, pn0.etapa, TRIM(pn0.dni)
      ORDER BY pn0.idpn DESC
    ) AS rn_pn
  FROM padronnominal pn0
  WHERE {" AND ".join(where)}
) pn
LEFT JOIN pdj ON pdj.ubigeo = pn.ubigeo AND pdj.periodo = DATE(pn.etapa)
LEFT JOIN pdrx pdr_cnv ON pdr_cnv.job_id = pdj.job_id AND pdr_cnv.cnv_key = TRIM(pn.dni)
LEFT JOIN pdrx pdr_dni ON pdr_dni.job_id = pdj.job_id AND pdr_dni.dni_key = TRIM(pn.dni)
LEFT JOIN pdrx pdr_cod ON pdr_cod.job_id = pdj.job_id AND pdr_cod.codpad_key = TRIM(pn.dni)
LEFT JOIN (
  SELECT
    ubigeo,
    etapa_mes,
    TRIM(dni_nino) AS dni_nino,
    MAX(CASE WHEN rn = 1 THEN fecha_intervencion END) AS primera_vd,
    MAX(CASE WHEN rn = 2 THEN fecha_intervencion END) AS segunda_vd,
    MAX(CASE WHEN rn = 3 THEN fecha_intervencion END) AS tercera_vd,
    MAX(CASE WHEN rn = 1 THEN etapa_text END) AS estadosvd,
    MAX(CASE WHEN rn = 2 THEN etapa_text END) AS estadosvd2,
    MAX(CASE WHEN rn = 3 THEN etapa_text END) AS estadosvd3
  FROM (
    SELECT
      y.ubigeo,
      y.etapa_mes,
      y.dni_nino,
      y.fecha_intervencion,
      y.etapa_text,
      ROW_NUMBER() OVER (
        PARTITION BY y.ubigeo, y.etapa_mes, y.dni_nino
        ORDER BY y.fecha_intervencion ASC
      ) AS rn
    FROM (
      SELECT
        d.ubigeo,
        d.etapa_mes,
        d.dni_nino,
        d.fecha_intervencion,
        d.etapa_text
      FROM (
        SELECT
          vr0.ubigeo,
          vr0.etapa_mes,
          TRIM(vr0.dni_nino) AS dni_nino,
          DATE(vr0.fecha_intervencion) AS fecha_intervencion,
          vr0.etapa_text,
          ROW_NUMBER() OVER (
            PARTITION BY vr0.ubigeo, vr0.etapa_mes, TRIM(vr0.dni_nino), DATE(vr0.fecha_intervencion)
            ORDER BY
              CASE
                WHEN LOWER(COALESCE(vr0.etapa_text,'')) LIKE 'visita%' THEN 3
                WHEN LOWER(COALESCE(vr0.etapa_text,'')) LIKE '%no encontrado%' THEN 2
                WHEN LOWER(COALESCE(vr0.etapa_text,'')) LIKE '%rechaz%' THEN 1
                ELSE 0
              END DESC,
              vr0.fecha_intervencion ASC
          ) AS rn_day
        FROM visitas_raw vr0
        WHERE vr0.etapa_mes IN ({','.join(['%s'] * len(etapas))})
          {vr_ubigeo_sql}
          AND vr0.fecha_intervencion IS NOT NULL
          AND (
            LOWER(COALESCE(vr0.etapa_text,'')) LIKE 'visita%'
            OR LOWER(COALESCE(vr0.etapa_text,'')) LIKE '%no encontrado%'
            OR LOWER(COALESCE(vr0.etapa_text,'')) LIKE '%rechaz%'
          )
      ) d
      WHERE d.rn_day = 1
    ) y
  ) x
  WHERE rn <= 3
  GROUP BY ubigeo, etapa_mes, TRIM(dni_nino)
) vr ON vr.ubigeo = pn.ubigeo AND vr.etapa_mes = pn.etapa AND vr.dni_nino = TRIM(pn.dni)
LEFT JOIN (
  SELECT
    dni,
    MAX(CASE WHEN rn = 1 THEN fecha_atencion END) AS fecha_atencion,
    MAX(CASE WHEN rn = 1 THEN hemoglobina END) AS hemoglobina,
    MAX(CASE WHEN rn = 1 THEN peso END) AS peso,
    MAX(CASE WHEN rn = 1 THEN talla END) AS talla,
    MAX(CASE WHEN rn = 1 THEN cie_10 END) AS cie_10,
    MAX(CASE WHEN rn = 1 THEN resultado END) AS resultado,
    MAX(CASE WHEN rn = 1 THEN hemoglobina END) AS hemoglobina1,
    MAX(CASE WHEN rn = 1 THEN fecha_atencion END) AS fecha_atencion1,
    MAX(CASE WHEN rn = 2 THEN hemoglobina END) AS hemoglobina2,
    MAX(CASE WHEN rn = 2 THEN fecha_atencion END) AS fecha_atencion2,
    MAX(CASE WHEN rn = 3 THEN hemoglobina END) AS hemoglobina3,
    MAX(CASE WHEN rn = 3 THEN fecha_atencion END) AS fecha_atencion3
  FROM (
    SELECT
      TRIM(rt.dni) AS dni,
      DATE(rt.fecha_atencion) AS fecha_atencion,
      rt.hemoglobina,
      rt.peso,
      rt.talla,
      rt.cie_10,
      rt.resultado,
      ROW_NUMBER() OVER (
        PARTITION BY TRIM(rt.dni)
        ORDER BY rt.fecha_atencion DESC, rt.id DESC
      ) AS rn
    FROM registro_tamizaje rt
    JOIN (
      SELECT DISTINCT TRIM(pn2.dni) AS dni
      FROM padronnominal pn2
      WHERE {" AND ".join(where_pn2)}
    ) dnis ON dnis.dni = TRIM(rt.dni)
    WHERE rt.fecha_atencion IS NOT NULL
  ) t
  WHERE rn <= 3
  GROUP BY dni
) tam ON tam.dni = TRIM(pn.dni)
LEFT JOIN (
  SELECT dni, nombrecompleto, apellidos
  FROM (
    SELECT TRIM(p.dni) AS dni, p.nombrecompleto, p.apellidos,
           ROW_NUMBER() OVER (PARTITION BY TRIM(p.dni) ORDER BY p.idpersona DESC) AS rn_persona
    FROM persona p
    WHERE TRIM(COALESCE(p.dni,'')) <> ''
  ) x
  WHERE rn_persona = 1
) p_actor ON p_actor.dni = TRIM(pn.actorsocial)
LEFT JOIN (
  SELECT dni, nombrecompleto, apellidos
  FROM (
    SELECT TRIM(p.dni) AS dni, p.nombrecompleto, p.apellidos,
           ROW_NUMBER() OVER (PARTITION BY TRIM(p.dni) ORDER BY p.idpersona DESC) AS rn_persona
    FROM persona p
    WHERE TRIM(COALESCE(p.dni,'')) <> ''
  ) x
  WHERE rn_persona = 1
) p_resp ON p_resp.dni = TRIM(pn.responsable)
LEFT JOIN ocurrencias o1 ON o1.idocurrencias = pn.idocurrencia
LEFT JOIN ocurrencias o2 ON o2.idocurrencias = pn.idocurrencia2
WHERE pn.rn_pn = 1
ORDER BY pn.ubigeo ASC, pn.etapa DESC, pn.idpn ASC
LIMIT {limit}
"""

    return sql, query_values


def label_for_col(name: str) -> str:
    n = (name or "").strip()
    m = {
        "idpn": "ID PN",
        "ubigeo": "Ubigeo",
        "etapa": "Etapa",
        "dni": "DNI",
        "fecha_nac": "F. Nac.",
        "dnimadre": "DNI Madre",
        "appatmadre": "Ap. Pat Madre",
        "apmatmadre": "Ap. Mat Madre",
        "nombresmadre": "Nombres Madre",
        "telefonopn": "Teléfono",
        "dni_padre": "DNI Padre",
        "nombre_padre": "Nombre Padre",
        "departamento": "Departamento",
        "provincia": "Provincia",
        "distrito": "Distrito",
        "iddistrito": "ID Distrito",
        "nuevadireccion": "Nueva dirección",
        "nuevareferencia": "Nueva referencia",
        "actorsocial": "Actor social (DNI)",
        "actor_nombre": "Actor social (Nombre)",
        "responsable": "Responsable (DNI)",
        "responsable_nombre": "Responsable (Nombre)",
        "idocurrencia": "ID Ocurrencia",
        "ocurrencia_desc": "Ocurrencia",
        "idocurrencia2": "ID Ocurrencia 2",
        "ocurrencia2_desc": "Ocurrencia 2",
        "estadovd": "Estado VD",
        "estadosvd": "EstadosVD",
        "estadosvd2": "EstadosVD2",
        "estadosvd3": "EstadosVD3",
        "nrovd": "Nro VD",
        "modovd": "Modo VD",
        "fechacita": "Fecha cita",
        "fecha_inicio_vd": "F. inicio VD",
        "fecha_fin_vd": "F. fin VD",
        "primera_vd": "1ra VD",
        "segunda_vd": "2da VD",
        "tercera_vd": "3ra VD",
        "programacion1": "Programación 1",
        "fechamodificacion": "F. modif.",
        "fechamodificacion2": "F. modif. 2",
        "estadointervencion": "Estado intervención",
        "tam_fecha_atencion": "F. atención (últ.)",
        "tam_hemoglobina": "Hemoglobina (últ.)",
        "tam_peso": "Peso (últ.)",
        "tam_talla": "Talla (últ.)",
        "tam_cie_10": "CIE10 (últ.)",
        "tam_resultado": "Resultado (últ.)",
        "hbregistro": "HB registro",
        "ccred": "CCRED",
        "estado_verificacion": "Estado verificación",
        "estado_verificado": "Estado verificado",
        "img_carnet": "Img carnet",
        "padronnominal": "Padrón nominal",
        "codigov": "Código V",
    }
    if n in m:
        return m[n]
    return n.replace("_", " ").strip().title()


def section_for_col(name: str) -> str:
    n = (name or "").strip()
    if n in (
        "idpn",
        "ubigeo",
        "etapa",
        "tipo",
        "rango",
        "tipodoc",
        "tipodocum",
        "dni",
        "nombres",
        "fecha_nac",
    ):
        return "Menor"
    if n in (
        "dnimadre",
        "appatmadre",
        "apmatmadre",
        "nombresmadre",
        "telefonopn",
        "telefono",
        "titular_linea",
        "celularseapp",
        "tiposeguro",
        "estadoseguro",
        "fecha_act_seguro",
    ):
        return "Madre"
    if n in ("dni_padre", "nombre_padre"):
        return "Padre"
    if n in (
        "departamento",
        "provincia",
        "distrito",
        "iddistrito",
        "ccpp",
        "zona",
        "mz",
        "direccion",
        "referencia",
        "nuevadireccion",
        "nuevareferencia",
        "lat",
        "lon",
        "lat2",
        "long2",
        "lat3",
        "long3",
        "eess_ua",
        "codeess",
        "nombre_comercial",
        "codqr",
        "codigov",
    ):
        return "Ubicación"
    if n in (
        "actorsocial",
        "actor_nombre",
        "responsable",
        "responsable_nombre",
        "usuario",
        "asignacion",
    ):
        return "Asignación"
    if n in ("idocurrencia", "ocurrencia_desc", "idocurrencia2", "ocurrencia2_desc"):
        return "Ocurrencias"
    if n in (
        "estadovd",
        "estadosvd",
        "estadosvd2",
        "estadosvd3",
        "nrovd",
        "modovd",
        "fechacita",
        "fecha_inicio_vd",
        "fecha_fin_vd",
        "primera_vd",
        "segunda_vd",
        "tercera_vd",
        "programacion1",
        "fechamodificacion",
        "fechamodificacion2",
        "estadointervencion",
    ):
        return "VD / Estados"
    return "Tamizaje / Otros"


def section_color(label: str) -> str:
    c = {
        "Menor": "E0F2FE",
        "Madre": "ECFDF5",
        "Padre": "FEF3C7",
        "Ubicación": "F3E8FF",
        "Asignación": "FFE4E6",
        "Ocurrencias": "E5E7EB",
        "VD / Estados": "DBEAFE",
        "Tamizaje / Otros": "FEE2E2",
    }
    return c.get(label, "E5E7EB")


def fmt_dmy(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        d = v.date()
    else:
        d = v
    try:
        y = int(getattr(d, "year", 0))
        m = int(getattr(d, "month", 0))
        dd = int(getattr(d, "day", 0))
        if y and m and dd:
            return f"{dd:02d}/{m:02d}/{y:04d}"
    except Exception:
        pass
    s = str(v).strip()
    if not s:
        return ""
    m = None
    try:
        m = datetime.fromisoformat(s.replace("Z", "")).date()
    except Exception:
        pass
    if m:
        return f"{m.day:02d}/{m.month:02d}/{m.year:04d}"
    return s


def fmt_dmy_hm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y %H:%M")
    s = str(v).strip()
    if not s:
        return ""
    return s


def make_cell(ws, value: Any, font: Font = None, fill: PatternFill = None, align: Alignment = None, border: Border = None, number_format: str = None):
    c = WriteOnlyCell(ws, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c


def to_date(v: Any):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        try:
            return datetime(int(v.year), int(v.month), int(v.day)).date()
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return None
    try:
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            return datetime.fromisoformat(s[:10]).date()
    except Exception:
        pass
    try:
        if len(s) >= 10 and s[2] == "/" and s[5] == "/":
            dd = int(s[0:2])
            mm = int(s[3:5])
            yy = int(s[6:10])
            return datetime(yy, mm, dd).date()
    except Exception:
        pass
    return None


def days_between(a: Any, b: Any):
    da = to_date(a)
    db = to_date(b)
    if not da or not db:
        return None
    return (db - da).days


def diff_age_parts(birth: Any, as_of: Any):
    b = to_date(birth)
    a = to_date(as_of)
    if not b or not a:
        return None
    if a < b:
        return None
    years = a.year - b.year
    months = a.month - b.month
    days = a.day - b.day
    if days < 0:
        prev_month_first = datetime(a.year, a.month, 1).date()
        prev_last_date = prev_month_first - timedelta(days=1)
        prev_last = prev_last_date.day
        days += prev_last
        months -= 1
    if months < 0:
        months += 12
        years -= 1
    total_days = (a - b).days
    return {"years": years, "months": months, "days": days, "total_days": total_days}


def is_no_encontrado(s: str) -> bool:
    v = (s or "").strip().lower()
    return "no encontrado" in v


def is_rechazado(s: str) -> bool:
    v = (s or "").strip().lower()
    return "rechaz" in v


def canon_estado_visita(s: str) -> str:
    if is_no_encontrado(s):
        return "No Encontrado"
    if is_rechazado(s):
        return "Rechazado"
    return (s or "").strip()


def calc_alerta_vd(getv):
    hoy = datetime.utcnow().date()
    primera = to_date(getv("primera_vd"))
    segunda = to_date(getv("segunda_vd"))
    tercera = to_date(getv("tercera_vd"))
    fechainicio_vd = to_date(getv("fecha_inicio_vd"))
    fechafin_vd = to_date(getv("fecha_fin_vd"))
    try:
        nrovd = int(getv("nrovd")) if getv("nrovd") is not None and str(getv("nrovd")).strip() != "" else None
    except Exception:
        nrovd = None
    try:
        tipo = int(getv("tipo")) if getv("tipo") is not None and str(getv("tipo")).strip() != "" else None
    except Exception:
        tipo = None

    limit_min = 13 if tipo == 6 else 7
    limit_max = 15 if tipo == 6 else 10
    limit_rgo = 15 if tipo == 6 else 9

    if not primera:
        nrovdr = 0
    elif primera and not segunda:
        nrovdr = 1
    elif primera and segunda and not tercera:
        nrovdr = 2
    else:
        nrovdr = 3

    estado1 = canon_estado_visita(str(getv("estadosvd") or ""))
    estado2 = canon_estado_visita(str(getv("estadosvd2") or ""))
    estado3 = canon_estado_visita(str(getv("estadosvd3") or ""))
    estado_actual = estado1 if nrovdr == 1 else estado2 if nrovdr == 2 else estado3 if nrovdr == 3 else ""
    if estado_actual and (is_no_encontrado(estado_actual) or is_rechazado(estado_actual)):
        return estado_actual

    if nrovdr == 0:
        if not fechainicio_vd:
            return ""
        d0 = days_between(fechainicio_vd, hoy)
        if d0 == 0:
            return "INICIO DE 1RA VD"
        if d0 is not None and d0 < 0:
            return "ESPERANDO FECHA DE 1RA VD"
        if d0 is not None and d0 > 5:
            return "CERO VD, FUERA DE FECHA"
        return "1RA VD EN CURSO"

    if nrovdr == 1:
        if not primera:
            return ""
        if (not segunda) and (isinstance(nrovd, int) and nrovd > 1):
            dt = days_between(primera, hoy)
            if dt is not None and dt > limit_max:
                return "INC. 2DA FECHA NO REALIZADA"
            ideal = primera + timedelta(days=limit_min)
            riesgo = primera + timedelta(days=limit_rgo)
            if ideal and hoy < ideal:
                return "ESPERANDO FECHA 2DA VD"
            if riesgo and hoy >= riesgo:
                return "VD EN RIESGO"
            return "2DA VISITA EN CURSO"
        if fechainicio_vd and primera < fechainicio_vd:
            return "INCONSISTENCIA VD ANTES DE INICIO DE INTERVENCION"
        if fechafin_vd and primera > fechafin_vd:
            return "INCONSISTENCIA VD FUERA DE MAXIMO PERMITIDO"
        if isinstance(nrovd, int) and nrovd == nrovdr:
            return "VD COMPLETO"
        return ""

    if nrovdr in (2, 3):
        if not primera or not segunda:
            return ""
        d12 = days_between(primera, segunda)
        if d12 is not None and (d12 < limit_min or d12 > limit_max):
            return "ERROR POR RANGO DE FECHA"
        if nrovdr == 3 and tercera:
            d23 = days_between(segunda, tercera)
            if d23 is not None and (d23 < limit_min or d23 > limit_max):
                return "ERROR POR RANGO DE FECHA"

        fechas = [primera, segunda] + ([tercera] if tercera else [])
        if fechainicio_vd and any(f < fechainicio_vd for f in fechas):
            return "INCONSISTENCIA VD ANTES DE FECHA PERMITIDO"
        if fechafin_vd and any(f > fechafin_vd for f in fechas):
            return "INCONSISTENCIA VD FUERA DE MAXIMO PERMITIDO"
        if isinstance(nrovd, int) and nrovdr > nrovd:
            return "SOBREPASO EL NRO DE VD PERMITIDO"
        if isinstance(nrovd, int) and nrovdr == nrovd:
            return "VD COMPLETO"
        return ""

    return ""


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--job_id", required=True, type=int)
    parser.add_argument("--out_path", required=True, type=str)
    args = parser.parse_args()

    load_dotenv(os.path.join(os.getcwd(), ".env.local"))

    job_id = int(args.job_id)
    out_path = str(args.out_path)

    db = connect_db()
    cur = db.cursor()
    try:
        cur.execute("SELECT params_json FROM report_export_jobs WHERE id=%s LIMIT 1", (job_id,))
        row = cur.fetchone()
        if not row:
            return 2
        params = json.loads(row[0] or "{}")
        tipo = str(params.get("tipo") or "").strip()
        etapas = normalize_etapas(params.get("etapas") or [])
        ubigeos = normalize_ubigeos(params.get("ubigeos") or [])
        if tipo not in ("1", "2") or not etapas:
            job_update(cur, job_id, "failed", 100, "Parámetros inválidos", None)
            db.commit()
            return 2

        os.makedirs(os.path.dirname(out_path), exist_ok=True)

        job_update(cur, job_id, "running", 1, "Preparando exportación", None)
        db.commit()
        log_line(f"Job {job_id}: iniciando")

        sql, values = build_query(tipo, etapas, ubigeos)
        job_update(cur, job_id, "running", 5, "Ejecutando consulta", None)
        db.commit()

        cur_stream = db.cursor(buffered=False)
        cur_stream.execute(sql, tuple(values))

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Padron")

        thin = Side(style="thin", color="D1D5DB")
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_meta_k = Font(bold=True)
        font_group = Font(bold=True, color="111827")
        font_header = Font(bold=True, color="FFFFFF")
        align_left = Alignment(horizontal="left", vertical="top", wrap_text=False)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill_header = PatternFill("solid", fgColor="111827")

        tipo_label = "Niños" if tipo == "1" else "Gestantes"
        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ubigeos_label = ",".join([str(x) for x in ubigeos]) if ubigeos else "TODOS"
        etapas_label = ",".join(etapas)

        row_idx = 0
        ws.append(
            [
                make_cell(ws, "Reporte:", font=font_meta_k, align=align_left),
                make_cell(ws, "Padrón nominal", align=align_left),
            ]
        )
        row_idx += 1
        ws.append([make_cell(ws, "Generado:", font=font_meta_k, align=align_left), make_cell(ws, generated_at, align=align_left)])
        row_idx += 1
        ws.append([make_cell(ws, "Tipo:", font=font_meta_k, align=align_left), make_cell(ws, f"{tipo_label} (tipovd={tipo})", align=align_left)])
        row_idx += 1
        ws.append([make_cell(ws, "Ubigeo:", font=font_meta_k, align=align_left), make_cell(ws, ubigeos_label, align=align_left)])
        row_idx += 1
        ws.append([make_cell(ws, "Etapas:", font=font_meta_k, align=align_left), make_cell(ws, etapas_label, align=align_left)])
        row_idx += 1
        ws.append([])
        row_idx += 1

        col_names = [d[0] for d in (cur_stream.description or [])]
        name_to_idx = {n: i for i, n in enumerate(col_names)}

        def gv(r, name: str):
            i = name_to_idx.get(name)
            if i is None:
                return None
            return r[i]

        def age_amd(r):
            age = diff_age_parts(gv(r, "fecha_nac"), gv(r, "etapa"))
            if not age:
                return ""
            return f"{age['years']}a {age['months']}m {age['days']}d"

        def age_days(r):
            age = diff_age_parts(gv(r, "fecha_nac"), gv(r, "etapa"))
            if not age:
                return ""
            return age["total_days"]

        def tel_madre(r):
            v = gv(r, "telefonopn")
            if v is None or str(v).strip() == "":
                v = gv(r, "telefono")
            return "" if v is None else str(v)

        def alerta_vd(r):
            return calc_alerta_vd(lambda n: gv(r, n))

        export_cols = [
            ("Menor", "N°", "raw", lambda r, idx: idx + 1),
            ("Menor", "ID PN", "raw", lambda r, _idx: gv(r, "idpn") or ""),
            ("Menor", "Ubigeo", "text", lambda r, _idx: gv(r, "ubigeo") or ""),
            ("Menor", "Etapa", "raw", lambda r, _idx: fmt_dmy(gv(r, "etapa"))),
            ("Menor", "Tipo", "raw", lambda r, _idx: gv(r, "tipo") or ""),
            ("Menor", "Rango", "raw", lambda r, _idx: gv(r, "rango") or ""),
            ("Menor", "Tipodoc", "raw", lambda r, _idx: gv(r, "tipodoc") or ""),
            ("Menor", "Tipodocum", "text", lambda r, _idx: gv(r, "tipodocum") or ""),
            ("Menor", "DNI", "text", lambda r, _idx: gv(r, "dni") or ""),
            ("Menor", "Nombres", "raw", lambda r, _idx: gv(r, "nombres") or ""),
            ("Menor", "F. Nac.", "raw", lambda r, _idx: fmt_dmy(gv(r, "fecha_nac"))),
            ("Menor", "Edad (a/m/d)", "raw", lambda r, _idx: age_amd(r)),
            ("Menor", "Edad (días)", "raw", lambda r, _idx: age_days(r)),
            ("Madre", "DNI Madre", "text", lambda r, _idx: gv(r, "dnimadre") or ""),
            ("Madre", "Ap. Pat Madre", "raw", lambda r, _idx: gv(r, "appatmadre") or ""),
            ("Madre", "Ap. Mat Madre", "raw", lambda r, _idx: gv(r, "apmatmadre") or ""),
            ("Madre", "Nombres Madre", "raw", lambda r, _idx: gv(r, "nombresmadre") or ""),
            ("Madre", "Teléfono", "text", lambda r, _idx: tel_madre(r)),
            ("Madre", "Titular línea", "raw", lambda r, _idx: gv(r, "titular_linea") or ""),
            ("Madre", "Celular SEAPP", "text", lambda r, _idx: gv(r, "celularseapp") or ""),
            ("Madre", "Tipo seguro", "raw", lambda r, _idx: gv(r, "tiposeguro") or ""),
            ("Madre", "Estado seguro", "raw", lambda r, _idx: gv(r, "estadoseguro") or ""),
            ("Madre", "Fecha act. seguro", "raw", lambda r, _idx: fmt_dmy_hm(gv(r, "fecha_act_seguro"))),
            ("Padre", "DNI Padre", "text", lambda r, _idx: gv(r, "dni_padre") or ""),
            ("Padre", "Nombre Padre", "raw", lambda r, _idx: gv(r, "nombre_padre") or ""),
            ("Ubicación", "Departamento", "raw", lambda r, _idx: gv(r, "departamento") or ""),
            ("Ubicación", "Provincia", "raw", lambda r, _idx: gv(r, "provincia") or ""),
            ("Ubicación", "Distrito", "raw", lambda r, _idx: gv(r, "distrito") or ""),
            ("Ubicación", "ID Distrito", "raw", lambda r, _idx: gv(r, "iddistrito") or ""),
            ("Ubicación", "CCPP", "raw", lambda r, _idx: gv(r, "ccpp") or ""),
            ("Ubicación", "Zona", "raw", lambda r, _idx: gv(r, "zona") or ""),
            ("Ubicación", "MZ", "raw", lambda r, _idx: gv(r, "mz") or ""),
            ("Ubicación", "Dirección", "raw", lambda r, _idx: gv(r, "direccion") or ""),
            ("Ubicación", "Referencia", "raw", lambda r, _idx: gv(r, "referencia") or ""),
            ("Ubicación", "Nueva dirección", "raw", lambda r, _idx: gv(r, "nuevadireccion") or ""),
            ("Ubicación", "Nueva referencia", "raw", lambda r, _idx: gv(r, "nuevareferencia") or ""),
            ("Ubicación", "Lat", "raw", lambda r, _idx: gv(r, "lat") or ""),
            ("Ubicación", "Lon", "raw", lambda r, _idx: gv(r, "lon") or ""),
            ("Ubicación", "Lat2", "raw", lambda r, _idx: gv(r, "lat2") or ""),
            ("Ubicación", "Long2", "raw", lambda r, _idx: gv(r, "long2") or ""),
            ("Ubicación", "Lat3", "raw", lambda r, _idx: gv(r, "lat3") or ""),
            ("Ubicación", "Long3", "raw", lambda r, _idx: gv(r, "long3") or ""),
            ("Ubicación", "EESS UA", "raw", lambda r, _idx: gv(r, "eess_ua") or ""),
            ("Ubicación", "CodEESS", "raw", lambda r, _idx: gv(r, "codeess") or ""),
            ("Ubicación", "Nombre comercial", "raw", lambda r, _idx: gv(r, "nombre_comercial") or ""),
            ("Ubicación", "CodQR", "raw", lambda r, _idx: gv(r, "codqr") or ""),
            ("Ubicación", "Código V", "raw", lambda r, _idx: gv(r, "codigov") or ""),
            ("Asignación", "Actor social (DNI)", "text", lambda r, _idx: gv(r, "actorsocial") or ""),
            ("Asignación", "Actor social (Nombre)", "raw", lambda r, _idx: gv(r, "actor_nombre") or ""),
            ("Asignación", "Responsable (DNI)", "text", lambda r, _idx: gv(r, "responsable") or ""),
            ("Asignación", "Responsable (Nombre)", "raw", lambda r, _idx: gv(r, "responsable_nombre") or ""),
            ("Asignación", "Usuario", "text", lambda r, _idx: gv(r, "usuario") or ""),
            ("Asignación", "Asignación", "raw", lambda r, _idx: gv(r, "asignacion") or ""),
            ("Ocurrencias", "ID Ocurrencia", "raw", lambda r, _idx: gv(r, "idocurrencia") or ""),
            ("Ocurrencias", "Ocurrencia", "raw", lambda r, _idx: gv(r, "ocurrencia_desc") or ""),
            ("Ocurrencias", "ID Ocurrencia 2", "raw", lambda r, _idx: gv(r, "idocurrencia2") or ""),
            ("Ocurrencias", "Ocurrencia 2", "raw", lambda r, _idx: gv(r, "ocurrencia2_desc") or ""),
            ("VD / Estados", "Estado VD", "raw", lambda r, _idx: gv(r, "estadovd") or ""),
            ("VD / Estados", "EstadosVD", "raw", lambda r, _idx: gv(r, "estadosvd") or ""),
            ("VD / Estados", "EstadosVD2", "raw", lambda r, _idx: gv(r, "estadosvd2") or ""),
            ("VD / Estados", "EstadosVD3", "raw", lambda r, _idx: gv(r, "estadosvd3") or ""),
            ("VD / Estados", "Nro VD", "raw", lambda r, _idx: gv(r, "nrovd") or ""),
            ("VD / Estados", "Modo VD", "raw", lambda r, _idx: gv(r, "modovd") or ""),
            ("VD / Estados", "Fecha cita", "raw", lambda r, _idx: fmt_dmy(gv(r, "fechacita"))),
            ("VD / Estados", "F. inicio VD", "raw", lambda r, _idx: fmt_dmy(gv(r, "fecha_inicio_vd"))),
            ("VD / Estados", "F. fin VD", "raw", lambda r, _idx: fmt_dmy(gv(r, "fecha_fin_vd"))),
            ("VD / Estados", "1ra VD", "raw", lambda r, _idx: fmt_dmy(gv(r, "primera_vd"))),
            ("VD / Estados", "2da VD", "raw", lambda r, _idx: fmt_dmy(gv(r, "segunda_vd"))),
            ("VD / Estados", "3ra VD", "raw", lambda r, _idx: fmt_dmy(gv(r, "tercera_vd"))),
            ("VD / Estados", "Días 1-2", "raw", lambda r, _idx: days_between(gv(r, "primera_vd"), gv(r, "segunda_vd")) or ""),
            ("VD / Estados", "Días 2-3", "raw", lambda r, _idx: days_between(gv(r, "segunda_vd"), gv(r, "tercera_vd")) or ""),
            ("VD / Estados", "Alerta VD", "raw", lambda r, _idx: alerta_vd(r)),
            ("VD / Estados", "Programación 1", "raw", lambda r, _idx: fmt_dmy(gv(r, "programacion1"))),
            ("VD / Estados", "F. modif.", "raw", lambda r, _idx: fmt_dmy_hm(gv(r, "fechamodificacion"))),
            ("VD / Estados", "F. modif. 2", "raw", lambda r, _idx: fmt_dmy_hm(gv(r, "fechamodificacion2"))),
            ("VD / Estados", "Estado intervención", "raw", lambda r, _idx: gv(r, "estadointervencion") or ""),
            ("Tamizaje / Otros", "Tipovd", "raw", lambda r, _idx: gv(r, "tipovd") or ""),
            ("Tamizaje / Otros", "Tamisaje", "raw", lambda r, _idx: gv(r, "tamisaje") or ""),
            ("Tamizaje / Otros", "Fecha tamisaje", "raw", lambda r, _idx: fmt_dmy(gv(r, "fechatamisaje"))),
            ("Tamizaje / Otros", "F. atención (últ.)", "raw", lambda r, _idx: fmt_dmy(gv(r, "tam_fecha_atencion"))),
            ("Tamizaje / Otros", "Hemoglobina (últ.)", "raw", lambda r, _idx: gv(r, "tam_hemoglobina") or ""),
            ("Tamizaje / Otros", "Peso (últ.)", "raw", lambda r, _idx: gv(r, "tam_peso") or ""),
            ("Tamizaje / Otros", "Talla (últ.)", "raw", lambda r, _idx: gv(r, "tam_talla") or ""),
            ("Tamizaje / Otros", "CIE10 (últ.)", "raw", lambda r, _idx: gv(r, "tam_cie_10") or ""),
            ("Tamizaje / Otros", "Resultado (últ.)", "raw", lambda r, _idx: gv(r, "tam_resultado") or ""),
            ("Tamizaje / Otros", "Hemoglobina1", "raw", lambda r, _idx: gv(r, "hemoglobina1") or ""),
            ("Tamizaje / Otros", "F. atención1", "raw", lambda r, _idx: fmt_dmy(gv(r, "fecha_atencion1"))),
            ("Tamizaje / Otros", "Hemoglobina2", "raw", lambda r, _idx: gv(r, "hemoglobina2") or ""),
            ("Tamizaje / Otros", "F. atención2", "raw", lambda r, _idx: fmt_dmy(gv(r, "fecha_atencion2"))),
            ("Tamizaje / Otros", "Hemoglobina3", "raw", lambda r, _idx: gv(r, "hemoglobina3") or ""),
            ("Tamizaje / Otros", "F. atención3", "raw", lambda r, _idx: fmt_dmy(gv(r, "fecha_atencion3"))),
            ("Tamizaje / Otros", "HB", "raw", lambda r, _idx: gv(r, "hb") or ""),
            ("Tamizaje / Otros", "Anemia", "raw", lambda r, _idx: gv(r, "anemia") or ""),
            ("Tamizaje / Otros", "Hierro", "raw", lambda r, _idx: gv(r, "hierro") or ""),
            ("Tamizaje / Otros", "TSF", "raw", lambda r, _idx: gv(r, "tsf") or ""),
            ("Tamizaje / Otros", "RSF", "raw", lambda r, _idx: gv(r, "rsf") or ""),
            ("Tamizaje / Otros", "Resultado", "raw", lambda r, _idx: gv(r, "resultado") or ""),
            ("Tamizaje / Otros", "Avance", "raw", lambda r, _idx: gv(r, "avance") or ""),
            ("Tamizaje / Otros", "HB registro", "raw", lambda r, _idx: gv(r, "hbregistro") or ""),
            ("Tamizaje / Otros", "CCRED", "raw", lambda r, _idx: gv(r, "ccred") or ""),
            ("Tamizaje / Otros", "Adulto", "raw", lambda r, _idx: gv(r, "adulto") or ""),
            ("Tamizaje / Otros", "Cantidad A", "raw", lambda r, _idx: gv(r, "cantidada") or ""),
            ("Tamizaje / Otros", "Discapacidad", "raw", lambda r, _idx: gv(r, "discapacidad") or ""),
            ("Tamizaje / Otros", "Visita DOPS", "raw", lambda r, _idx: gv(r, "visitadops") or ""),
            ("Tamizaje / Otros", "Sesión DEM", "raw", lambda r, _idx: gv(r, "sesiondem") or ""),
            ("Tamizaje / Otros", "Tiene PS", "raw", lambda r, _idx: gv(r, "tieneps") or ""),
            ("Tamizaje / Otros", "Observación", "raw", lambda r, _idx: gv(r, "observacion") or ""),
            ("Tamizaje / Otros", "Observación 2", "raw", lambda r, _idx: gv(r, "observacion2") or ""),
            ("Tamizaje / Otros", "Obs padrón", "raw", lambda r, _idx: gv(r, "obspadron") or ""),
            ("Tamizaje / Otros", "Estado verificación", "raw", lambda r, _idx: gv(r, "estado_verificacion") or ""),
            ("Tamizaje / Otros", "Estado", "raw", lambda r, _idx: gv(r, "estado") or ""),
            ("Tamizaje / Otros", "Estado verificado", "raw", lambda r, _idx: gv(r, "estado_verificado") or ""),
            ("Tamizaje / Otros", "Tipo dispositivo", "raw", lambda r, _idx: gv(r, "tipodispositivo") or ""),
            ("Tamizaje / Otros", "Fotos", "raw", lambda r, _idx: gv(r, "fotos") or ""),
            ("Tamizaje / Otros", "Img carnet", "raw", lambda r, _idx: gv(r, "img_carnet") or ""),
            ("Tamizaje / Otros", "Padrón nominal", "raw", lambda r, _idx: gv(r, "padronnominal") or ""),
        ]

        out_labels = [c[1] for c in export_cols]
        out_sections = [c[0] for c in export_cols]

        group_row_num = row_idx + 1
        group_cells = []
        spans = []
        span_start = 1
        span_label = out_sections[0]
        for i, s in enumerate(out_sections, start=1):
            if i == 1:
                span_start = 1
                span_label = s
            if s != span_label:
                spans.append((span_start, i - 1, span_label))
                span_start = i
                span_label = s
        spans.append((span_start, len(out_sections), span_label))

        span_starts = {a: lab for a, _b, lab in spans}

        for i, s in enumerate(out_sections, start=1):
            if i in span_starts:
                fill_group = PatternFill("solid", fgColor=section_color(s))
                group_cells.append(make_cell(ws, s, font=font_group, fill=fill_group, align=align_center, border=border_thin))
            else:
                fill_group = PatternFill("solid", fgColor=section_color(s))
                group_cells.append(make_cell(ws, "", font=font_group, fill=fill_group, align=align_center, border=border_thin))
        ws.append(group_cells)
        row_idx += 1

        try:
            for a, b, _lab in spans:
                if b > a:
                    ws.merge_cells(start_row=group_row_num, start_column=a, end_row=group_row_num, end_column=b)
        except Exception:
            pass

        header_row_num = row_idx + 1
        ws.append(
            [
                make_cell(ws, lab, font=font_header, fill=fill_header, align=align_left, border=border_thin)
                for lab in out_labels
            ]
        )
        row_idx += 1
        try:
            ws.freeze_panes = f"A{header_row_num + 1}"
        except Exception:
            pass

        text_cols = {
            "ubigeo",
            "dni",
            "dnimadre",
            "dni_padre",
            "telefono",
            "telefonopn",
            "celularseapp",
            "codqr",
            "codigov",
            "actorsocial",
            "responsable",
            "tipodocum",
        }
        date_cols = {
            "etapa",
            "fecha_nac",
            "fechacita",
            "fecha_inicio_vd",
            "fecha_fin_vd",
            "primera_vd",
            "segunda_vd",
            "tercera_vd",
            "programacion1",
            "fechatamisaje",
            "tam_fecha_atencion",
            "fecha_atencion1",
            "fecha_atencion2",
            "fecha_atencion3",
        }
        datetime_cols = {"fechamodificacion", "fechamodificacion2", "fecha_act_seguro"}

        last_update = 0
        written = 0
        while True:
            batch = cur_stream.fetchmany(500)
            if not batch:
                break
            for r in batch:
                out_row = []
                for sec, lab, kind, fn in export_cols:
                    v = fn(r, written)
                    if kind == "text":
                        out_row.append(make_cell(ws, "" if v is None else str(v), number_format="@"))
                    else:
                        out_row.append("" if v is None else v)
                ws.append(out_row)
                written += 1
            if written - last_update >= 5000:
                last_update = written
                job_update(cur, job_id, "running", 10, f"Escribiendo filas: {written}", None)
                db.commit()
                log_line(f"Job {job_id}: filas={written}")
        last_row = header_row_num + written
        last_col_letter = get_column_letter(len(out_labels))
        try:
            ws.auto_filter.ref = f"A{header_row_num}:{last_col_letter}{last_row}"
        except Exception:
            pass

        try:
            for idx, (_sec, lab, kind, _fn) in enumerate(export_cols, start=1):
                letter = get_column_letter(idx)
                w = 14
                if lab in ("N°",):
                    w = 6
                if "Nombres" in lab or "Dirección" in lab or "Referencia" in lab or "Observación" in lab or "Nombre Padre" in lab:
                    w = 30
                if "Observación" in lab:
                    w = 40
                if kind == "text":
                    w = 16
                ws.column_dimensions[letter].width = w
        except Exception:
            pass


        wb.save(out_path)
        job_update(cur, job_id, "done", 100, f"Listo. Filas: {written}", out_path)
        db.commit()
        log_line(f"Job {job_id}: terminado filas={written}")
    except Exception as e:
        try:
            job_update(cur, job_id, "failed", 100, f"Error: {str(e)[:450]}", None)
            db.commit()
        except Exception:
            pass
        log_line(f"Job {job_id}: error {e}")
        return 1
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


if __name__ == "__main__":
    sys.exit(main())
