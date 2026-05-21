import argparse
import os
import re
import sys
import time
from datetime import date, datetime, timedelta

import mysql.connector
from dotenv import load_dotenv
from openpyxl import load_workbook


def log_line(msg: str):
    p = os.getenv("PADRON_VD_LOG_PATH", "").strip()
    if not p:
        return
    try:
        with open(p, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.utcnow().isoformat(timespec='seconds')}Z] {msg}\n")
    except Exception:
        pass


def connect_db():
    host = os.getenv("DB_HOST", "127.0.0.1")
    port = int(os.getenv("DB_PORT", "3306"))
    user = os.getenv("DB_USER", "")
    password = os.getenv("DB_PASSWORD", "")
    database = os.getenv("DB_NAME", "")
    return mysql.connector.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        autocommit=False,
    )


def to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_int(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if v != v:
            return None
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d\-]+", "", s)
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return None
        base = date(1899, 12, 30)
        try:
            return base + timedelta(days=int(v))
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def job_update(cur, job_id: str, **fields):
    if not fields:
        return
    set_parts = []
    vals = []
    for k, v in fields.items():
        set_parts.append(f"{k} = %s")
        vals.append(v)
    vals.append(job_id)
    cur.execute(f"UPDATE padron_vd_import_jobs SET {', '.join(set_parts)} WHERE id = %s", vals)


def load_config(cur):
    cur.execute("SELECT * FROM padron_vd_import_config WHERE id = 1 LIMIT 1")
    r = cur.fetchone()
    if not r:
        raise Exception("No existe configuración de columnas (padron_vd_import_config).")
    cols = [d[0] for d in cur.description]
    return dict(zip(cols, r))


LAST_FIELDS = [
    "dni",
    "ccpp",
    "referencia",
    "idocurrencia",
    "idocurrencia2",
    "obspadron",
    "fechacita",
    "departamento",
    "provincia",
    "distrito",
]


def fetch_last_by_dni(cur, dni_list):
    out = {}
    if not dni_list:
        return out
    chunk = 900
    for i in range(0, len(dni_list), chunk):
        part = dni_list[i : i + chunk]
        placeholders = ",".join(["%s"] * len(part))
        sql = f"""
          SELECT pn.dni, pn.ccpp, pn.referencia, pn.idocurrencia, pn.idocurrencia2, pn.obspadron,
                 pn.fechacita, pn.departamento, pn.provincia, pn.distrito
          FROM padronnominal pn
          JOIN (
            SELECT dni, MAX(idpn) AS idpn
            FROM padronnominal
            WHERE tipovd <> '4' AND dni IN ({placeholders})
            GROUP BY dni
          ) t ON t.idpn = pn.idpn
        """
        cur.execute(sql, part)
        for row in cur.fetchall():
            rec = dict(zip(LAST_FIELDS, row))
            out[str(rec.get("dni") or "").strip()] = rec
    return out


def fetch_existing_dni_by_ubigeo_etapa(cur, ubigeo: int, etapa_val: date, dni_list):
    out = set()
    if not dni_list:
        return out
    chunk = 900
    for i in range(0, len(dni_list), chunk):
        part = dni_list[i : i + chunk]
        placeholders = ",".join(["%s"] * len(part))
        sql = f"""
          SELECT dni
          FROM padronnominal
          WHERE ubigeo = %s
            AND DATE_FORMAT(etapa, '%%Y-%%m-01') = %s
            AND TRIM(COALESCE(tipovd,'')) = '1'
            AND dni IN ({placeholders})
        """
        cur.execute(sql, [ubigeo, etapa_val] + part)
        for row in cur.fetchall():
            dni = str(row[0] or "").strip()
            if dni:
                out.add(dni)
    return out


INSERT_SQL = """
INSERT INTO padronnominal (
  rango, ccpp, direccion, dni, nombres, fecha_nac, dnimadre,
  telefono, etapa, nrovd, fecha_inicio_vd, fecha_fin_vd, ubigeo,
  tipovd, actorsocial, responsable, eess_ua,
  idocurrencia, idocurrencia2, obspadron, fechacita,
  departamento, provincia, distrito, referencia
) VALUES (
  %s,%s,%s,%s,%s,%s,%s,
  %s,%s,%s,%s,%s,%s,
  %s,%s,%s,%s,
  %s,%s,%s,%s,
  %s,%s,%s,%s
)
"""


def get_cell(row_vals, idx):
    if idx is None:
        return None
    if idx < 0:
        return None
    if idx >= len(row_vals):
        return None
    return row_vals[idx]


def etapa_from_month(d: date):
    return date(d.year, d.month, 1)


def run_import(job_id: str, file_path: str):
    load_dotenv(os.getenv("PADRON_VD_DOTENV", ".env.local"), override=False)
    load_dotenv(override=False)

    log_line(f"Job {job_id} starting. File: {file_path}")
    db = connect_db()
    cur = db.cursor()

    cfg = load_config(cur)
    db.commit()

    job_update(
        cur,
        job_id,
        status="running",
        progress=0,
        started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        message="Leyendo Excel...",
    )
    db.commit()

    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet_index = int(cfg.get("sheet_index") or 0)
    if sheet_index < 0 or sheet_index >= len(wb.worksheets):
        sheet_index = 0
    ws = wb.worksheets[sheet_index]

    start_row = int(cfg.get("start_row") or 1)
    if start_row < 1:
        start_row = 1

    records = []
    for row_vals in ws.iter_rows(min_row=start_row, values_only=True):
        dni = to_text(get_cell(row_vals, cfg.get("col_dni")))
        if not dni:
            continue
        ubigeo = to_int(get_cell(row_vals, cfg.get("col_ubigeo")))
        fecha_nac = to_date(get_cell(row_vals, cfg.get("col_fecha_nac")))

        records.append(
            {
                "dni": dni,
                "ubigeo": ubigeo,
                "fecha_nac": fecha_nac,
                "rango": to_text(get_cell(row_vals, cfg.get("col_rango"))),
                "ccpp": to_text(get_cell(row_vals, cfg.get("col_ccpp"))),
                "direccion": to_text(get_cell(row_vals, cfg.get("col_direccion"))),
                "dnimadre": to_text(get_cell(row_vals, cfg.get("col_dnimadre"))),
                "telefono": to_text(get_cell(row_vals, cfg.get("col_telefono"))),
                "actorsocial": to_text(get_cell(row_vals, cfg.get("col_actorsocial"))),
                "responsable": to_text(get_cell(row_vals, cfg.get("col_responsable"))),
                "departamento": to_text(get_cell(row_vals, cfg.get("col_departamento"))),
                "provincia": to_text(get_cell(row_vals, cfg.get("col_provincia"))),
                "distrito": to_text(get_cell(row_vals, cfg.get("col_distrito"))),
                "eess_ua": to_text(get_cell(row_vals, cfg.get("col_eess_ua"))),
                "nrovd": to_int(get_cell(row_vals, cfg.get("col_nrovd"))),
                "fecha_inicio_vd": to_date(get_cell(row_vals, cfg.get("col_fecha_inicio_vd"))),
                "fecha_fin_vd": to_date(get_cell(row_vals, cfg.get("col_fecha_fin_vd"))),
                "etapa": to_date(get_cell(row_vals, cfg.get("col_etapa"))),
            }
        )

    total_rows = len(records)
    if total_rows == 0:
        job_update(
            cur,
            job_id,
            status="failed",
            finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            message="No se encontraron registros (DNI vacío). Revisa fila inicio y columnas.",
        )
        db.commit()
        return

    job_update(cur, job_id, total_rows=total_rows, processed_rows=0, inserted_rows=0, progress=0)
    db.commit()

    dni_unique = sorted(set([r["dni"] for r in records if r.get("dni")]))
    log_line(f"Found rows={total_rows} unique_dni={len(dni_unique)}")

    last_map = fetch_last_by_dni(cur, dni_unique)
    db.commit()

    batch_size = 500
    inserted_total = 0
    skipped_total = 0
    processed = 0
    last_tick = time.time()

    cur.execute("SET SESSION sql_mode = ''")

    for i in range(0, total_rows, batch_size):
        batch = records[i : i + batch_size]
        values = []
        computed = []
        groups = {}

        for r in batch:
            dni = str(r.get("dni") or "").strip()
            if not dni:
                continue
            ubigeo = r.get("ubigeo")
            if ubigeo is None:
                continue
            ubigeo_i = int(ubigeo)

            etapa_val = r.get("etapa")
            if etapa_val is None:
                fi = r.get("fecha_inicio_vd")
                if fi is not None:
                    etapa_val = etapa_from_month(fi)
            if etapa_val is None:
                continue
            if isinstance(etapa_val, datetime):
                etapa_val = etapa_val.date()
            if isinstance(etapa_val, date):
                etapa_val = etapa_from_month(etapa_val)

            k = (ubigeo_i, etapa_val)
            groups.setdefault(k, set()).add(dni)
            computed.append((r, ubigeo_i, etapa_val, dni))

        existing_by_key = {}
        for k, dset in groups.items():
            u, e = k
            existing_by_key[k] = fetch_existing_dni_by_ubigeo_etapa(cur, u, e, sorted(dset))

        for r, ubigeo_i, etapa_val, dni in computed:
            if dni in existing_by_key.get((ubigeo_i, etapa_val), set()):
                skipped_total += 1
                continue

            last = last_map.get(dni, {})
            ccpp_raw = r.get("ccpp")
            ccpp_excel = to_text(ccpp_raw)
            if not ccpp_excel or ccpp_excel.strip() == "#N/D":
                ccpp = to_text(last.get("ccpp"))
                referencia = to_text(last.get("referencia")) or ""
            else:
                ccpp = ccpp_excel
                referencia = ""

            dep = r.get("departamento") or to_text(last.get("departamento")) or ""
            prov = r.get("provincia") or to_text(last.get("provincia")) or ""
            dist = r.get("distrito") or to_text(last.get("distrito")) or ""

            last_ido = to_int(last.get("idocurrencia")) or 0
            last_ido2 = to_int(last.get("idocurrencia2")) or last_ido

            values.append(
                (
                    r.get("rango") or None,
                    ccpp or None,
                    r.get("direccion") or None,
                    dni,
                    "NULL",
                    r.get("fecha_nac") or None,
                    r.get("dnimadre") or None,
                    r.get("telefono") or None,
                    etapa_val,
                    r.get("nrovd") or None,
                    r.get("fecha_inicio_vd") or None,
                    r.get("fecha_fin_vd") or None,
                    ubigeo_i,
                    "1",
                    r.get("actorsocial") or None,
                    r.get("responsable") or None,
                    r.get("eess_ua") or None,
                    last_ido,
                    last_ido2,
                    to_text(last.get("obspadron")) or "",
                    to_date(last.get("fechacita")),
                    dep,
                    prov,
                    dist,
                    referencia,
                )
            )

        if values:
            cur.executemany(INSERT_SQL, values)
            db.commit()
            inserted_total += cur.rowcount if cur.rowcount > 0 else 0

        processed += len(batch)
        now = time.time()
        if now - last_tick >= 0.6 or processed >= total_rows:
            progress = int((processed / max(1, total_rows)) * 100)
            job_update(
                cur,
                job_id,
                progress=progress,
                processed_rows=processed,
                inserted_rows=inserted_total,
                message=f"Procesando... {processed}/{total_rows} · Omitidos: {skipped_total}",
            )
            db.commit()
            last_tick = now

    job_update(
        cur,
        job_id,
        status="done",
        progress=100,
        processed_rows=processed,
        inserted_rows=inserted_total,
        finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        message=f"Completado. Insertados: {inserted_total} · Omitidos: {skipped_total}",
    )
    db.commit()
    log_line(f"Done. processed={processed} inserted={inserted_total} skipped={skipped_total}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--job", required=True)
    ap.add_argument("--file", required=True)
    args = ap.parse_args()

    try:
        run_import(args.job, args.file)
    except Exception as e:
        try:
            load_dotenv(os.getenv("PADRON_VD_DOTENV", ".env.local"), override=False)
            load_dotenv(override=False)
            db = connect_db()
            cur = db.cursor()
            job_update(
                cur,
                args.job,
                status="failed",
                progress=0,
                finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                message=str(e),
            )
            db.commit()
        except Exception:
            pass
        log_line(f"FAILED: {e}")
        raise


if __name__ == "__main__":
    main()

