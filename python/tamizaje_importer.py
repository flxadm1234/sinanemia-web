import argparse
import os
import re
import sys
import time
from datetime import date, datetime

import mysql.connector
from dotenv import load_dotenv
from openpyxl import load_workbook


REQUIRED_HEADERS = [
    "Id_Cita",
    "Lote",
    "UPS",
    "NOMBRE_PERSONAL",
    "Nombre_Registrador",
    "periodo",
    "renaes",
    "Red",
    "MicroRed",
    "Provincia",
    "Distrito",
    "Tipo_documento",
    "dni",
    "sexo",
    "fecha_nacimiento",
    "Fecha_Atencion",
    "PESO",
    "TALLA",
    "HEMOGLOBINA",
    "gruporiesgo_desc",
    "condicion_gestante",
    "Tipo_Edad_PAC",
    "ANIO_ACTUAL_PAC",
    "MES_ACTUAL_PAC",
    "DIA_ACTUAL_PAC",
    "Nombre_Establecimiento",
    "CIE_10",
    "Diagnostico",
    "LAB1",
    "LAB2",
    "LAB3",
    "RESULTADO",
    "TOTAL",
]


def norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[\s_]+", "", s)
    s = s.replace("ó", "o").replace("í", "i").replace("á", "a").replace("é", "e").replace("ú", "u")
    return s


REQUIRED_NORM = [norm_header(h) for h in REQUIRED_HEADERS]
REQUIRED_MAP = {norm_header(h): h for h in REQUIRED_HEADERS}

def log_line(msg: str):
    p = os.getenv("TAMIZAJE_LOG_PATH", "").strip()
    if not p:
        return
    try:
        with open(p, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.utcnow().isoformat(timespec='seconds')}Z] {msg}\n")
    except Exception:
        pass


def to_int(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, float):
        if v != v:
            return None
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d\-]+", "", s)
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def to_decimal(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return None
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d = int(m.group(1))
        mo = int(m.group(2))
        y = int(m.group(3))
        try:
            return date(y, mo, d)
        except Exception:
            return None
    m2 = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m2:
        y = int(m2.group(1))
        mo = int(m2.group(2))
        d = int(m2.group(3))
        try:
            return date(y, mo, d)
        except Exception:
            return None
    return None


def ensure_tables(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS tamizaje_import_jobs (
          id CHAR(36) PRIMARY KEY,
          status ENUM('queued','running','done','failed') NOT NULL,
          progress INT NOT NULL DEFAULT 0,
          total_rows INT NOT NULL DEFAULT 0,
          processed_rows INT NOT NULL DEFAULT 0,
          inserted_rows INT NOT NULL DEFAULT 0,
          file_name VARCHAR(255) NULL,
          source VARCHAR(30) NOT NULL DEFAULT 'web',
          requested_by VARCHAR(15) NULL,
          started_at DATETIME NULL,
          finished_at DATETIME NULL,
          message TEXT NULL,
          created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
          updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS registro_tamizaje (
          id INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
          id_cita BIGINT NULL,
          lote VARCHAR(20) NULL,
          ups VARCHAR(200) NULL,
          nombre_personal VARCHAR(200) NULL,
          nombre_registrador VARCHAR(200) NULL,
          periodo INT NULL,
          renaes VARCHAR(20) NULL,
          red VARCHAR(200) NULL,
          microred VARCHAR(200) NULL,
          provincia VARCHAR(200) NULL,
          distrito VARCHAR(200) NULL,
          tipo_documento VARCHAR(40) NULL,
          dni VARCHAR(20) NULL,
          sexo VARCHAR(5) NULL,
          fecha_nacimiento DATE NULL,
          fecha_atencion DATE NULL,
          peso DECIMAL(8,2) NULL,
          talla DECIMAL(8,2) NULL,
          hemoglobina DECIMAL(8,2) NULL,
          grupo_riesgo_desc VARCHAR(250) NULL,
          condicion_gestante VARCHAR(250) NULL,
          tipo_edad_pac VARCHAR(10) NULL,
          anio_actual_pac INT NULL,
          mes_actual_pac INT NULL,
          dia_actual_pac INT NULL,
          nombre_establecimiento VARCHAR(250) NULL,
          cie_10 VARCHAR(40) NULL,
          diagnostico VARCHAR(300) NULL,
          lab1 VARCHAR(40) NULL,
          lab2 VARCHAR(40) NULL,
          lab3 VARCHAR(40) NULL,
          resultado VARCHAR(200) NULL,
          total INT NULL,
          created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
          INDEX idx_dni (dni),
          INDEX idx_fecha_atencion (fecha_atencion),
          INDEX idx_periodo (periodo),
          INDEX idx_renaes (renaes),
          INDEX idx_id_cita (id_cita),
          INDEX idx_dni_fecha (dni, fecha_atencion)
        ) ENGINE=InnoDB
        """
    )


def job_update(cur, job_id: str, **fields):
    keys = list(fields.keys())
    if not keys:
        return
    set_sql = ", ".join([f"{k}=%s" for k in keys])
    vals = [fields[k] for k in keys]
    vals.append(job_id)
    cur.execute(f"UPDATE tamizaje_import_jobs SET {set_sql} WHERE id=%s", vals)


def find_header_row(ws):
    max_scan = 40
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        values = list(row) if row is not None else []
        norms = [norm_header(str(v)) if v is not None else "" for v in values]
        present = set([n for n in norms if n])
        hits = sum(1 for req in REQUIRED_NORM if req in present)
        if hits >= int(len(REQUIRED_NORM) * 0.7) and REQUIRED_NORM[0] in present:
            return i, norms, len(values)
    return None, None, None


def build_col_map(header_norms):
    idx = {}
    for i, hn in enumerate(header_norms):
        if not hn:
            continue
        idx.setdefault(hn, i)
    missing = [h for h in REQUIRED_NORM if h not in idx]
    return idx, missing


def connect_db():
    host = os.getenv("DB_HOST", "127.0.0.1")
    port = int(os.getenv("DB_PORT", "3306"))
    user = os.getenv("DB_USER", "")
    password = os.getenv("DB_PASSWORD", "")
    database = os.getenv("DB_NAME", "")
    return mysql.connector.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        autocommit=False,
    )


INSERT_SQL = """
INSERT INTO registro_tamizaje (
  id_cita, lote, ups, nombre_personal, nombre_registrador, periodo,
  renaes, red, microred, provincia, distrito,
  tipo_documento, dni, sexo, fecha_nacimiento, fecha_atencion,
  peso, talla, hemoglobina, grupo_riesgo_desc, condicion_gestante,
  tipo_edad_pac, anio_actual_pac, mes_actual_pac, dia_actual_pac,
  nombre_establecimiento, cie_10, diagnostico,
  lab1, lab2, lab3, resultado, total
) VALUES (
  %s,%s,%s,%s,%s,%s,
  %s,%s,%s,%s,%s,
  %s,%s,%s,%s,%s,
  %s,%s,%s,%s,%s,
  %s,%s,%s,%s,
  %s,%s,%s,
  %s,%s,%s,%s,%s
)
"""


def get_cell(row_vals, col_map, key_norm):
    i = col_map.get(key_norm)
    if i is None or i >= len(row_vals):
        return None
    return row_vals[i]


def run_import(job_id: str, file_path: str):
    load_dotenv(os.getenv("TAMIZAJE_DOTENV", ".env.local"), override=False)
    load_dotenv(override=False)

    log_line(f"Job {job_id} starting. File: {file_path}")
    db = connect_db()
    cur = db.cursor()
    ensure_tables(cur)
    db.commit()

    job_update(
        cur,
        job_id,
        status="running",
        progress=0,
        started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        message="Iniciando importación...",
    )
    db.commit()

    log_line("Opening workbook (read_only=True, data_only=True)")
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_row, header_norms, header_len = find_header_row(ws)
    if not header_row:
        log_line("Header row not found")
        job_update(cur, job_id, status="failed", progress=0, finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"), message="No se encontró el encabezado. La plantilla no coincide.")
        db.commit()
        return 2

    col_map, missing = build_col_map(header_norms)
    if missing:
        log_line("Missing columns: " + ", ".join([REQUIRED_MAP.get(m, m) for m in missing[:12]]))
        job_update(
            cur,
            job_id,
            status="failed",
            progress=0,
            finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            message="Plantilla incorrecta. Faltan columnas: "
            + ", ".join([REQUIRED_MAP.get(m, m) for m in missing[:12]]),
        )
        db.commit()
        return 2

    total_rows = max(0, (ws.max_row or 0) - int(header_row))
    if total_rows <= 0:
        log_line("No data rows detected")
        job_update(
            cur,
            job_id,
            status="failed",
            progress=0,
            finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            message="El archivo no contiene filas de datos.",
        )
        db.commit()
        return 2

    job_update(cur, job_id, total_rows=total_rows, message="Limpiando tabla y cargando datos...")
    db.commit()

    log_line(f"TRUNCATE registro_tamizaje, total_rows={total_rows}")
    cur.execute("TRUNCATE TABLE registro_tamizaje")
    db.commit()

    batch = []
    scanned = 0
    inserted = 0
    last_update = time.time()

    def flush():
        nonlocal inserted
        if not batch:
            return
        cur.executemany(INSERT_SQL, batch)
        inserted += cur.rowcount if cur.rowcount is not None else 0
        db.commit()
        batch.clear()

    idx_id_cita = col_map.get(norm_header("Id_Cita"))
    idx_lote = col_map.get(norm_header("Lote"))
    idx_ups = col_map.get(norm_header("UPS"))
    idx_nombre_personal = col_map.get(norm_header("NOMBRE_PERSONAL"))
    idx_nombre_registrador = col_map.get(norm_header("Nombre_Registrador"))
    idx_periodo = col_map.get(norm_header("periodo"))
    idx_renaes = col_map.get(norm_header("renaes"))
    idx_red = col_map.get(norm_header("Red"))
    idx_microred = col_map.get(norm_header("MicroRed"))
    idx_provincia = col_map.get(norm_header("Provincia"))
    idx_distrito = col_map.get(norm_header("Distrito"))
    idx_tipo_documento = col_map.get(norm_header("Tipo_documento"))
    idx_dni = col_map.get(norm_header("dni"))
    idx_sexo = col_map.get(norm_header("sexo"))
    idx_fecha_nacimiento = col_map.get(norm_header("fecha_nacimiento"))
    idx_fecha_atencion = col_map.get(norm_header("Fecha_Atencion"))
    idx_peso = col_map.get(norm_header("PESO"))
    idx_talla = col_map.get(norm_header("TALLA"))
    idx_hemoglobina = col_map.get(norm_header("HEMOGLOBINA"))
    idx_gruporiesgo_desc = col_map.get(norm_header("gruporiesgo_desc"))
    idx_condicion_gestante = col_map.get(norm_header("condicion_gestante"))
    idx_tipo_edad_pac = col_map.get(norm_header("Tipo_Edad_PAC"))
    idx_anio_actual_pac = col_map.get(norm_header("ANIO_ACTUAL_PAC"))
    idx_mes_actual_pac = col_map.get(norm_header("MES_ACTUAL_PAC"))
    idx_dia_actual_pac = col_map.get(norm_header("DIA_ACTUAL_PAC"))
    idx_nombre_establecimiento = col_map.get(norm_header("Nombre_Establecimiento"))
    idx_cie_10 = col_map.get(norm_header("CIE_10"))
    idx_diagnostico = col_map.get(norm_header("Diagnostico"))
    idx_lab1 = col_map.get(norm_header("LAB1"))
    idx_lab2 = col_map.get(norm_header("LAB2"))
    idx_lab3 = col_map.get(norm_header("LAB3"))
    idx_resultado = col_map.get(norm_header("RESULTADO"))
    idx_total = col_map.get(norm_header("TOTAL"))

    def at(row_vals, idx):
        if idx is None:
            return None
        if idx < 0 or idx >= len(row_vals):
            return None
        return row_vals[idx]

    empty_streak = 0
    max_col = int(header_len or 0) if header_len else None
    for row_vals in ws.iter_rows(min_row=header_row + 1, values_only=True, max_col=max_col):
        if row_vals is None:
            continue
        scanned += 1
        id_cita = to_int(at(row_vals, idx_id_cita))
        if id_cita is None:
            if not any(v is not None and str(v).strip() != "" for v in row_vals):
                empty_streak += 1
                if empty_streak >= 2000:
                    break
            continue
        empty_streak = 0

        rec = (
            id_cita,
            to_text(at(row_vals, idx_lote)),
            to_text(at(row_vals, idx_ups)),
            to_text(at(row_vals, idx_nombre_personal)),
            to_text(at(row_vals, idx_nombre_registrador)),
            to_int(at(row_vals, idx_periodo)),
            to_text(at(row_vals, idx_renaes)),
            to_text(at(row_vals, idx_red)),
            to_text(at(row_vals, idx_microred)),
            to_text(at(row_vals, idx_provincia)),
            to_text(at(row_vals, idx_distrito)),
            to_text(at(row_vals, idx_tipo_documento)),
            to_text(at(row_vals, idx_dni)),
            to_text(at(row_vals, idx_sexo)),
            to_date(at(row_vals, idx_fecha_nacimiento)),
            to_date(at(row_vals, idx_fecha_atencion)),
            to_decimal(at(row_vals, idx_peso)),
            to_decimal(at(row_vals, idx_talla)),
            to_decimal(at(row_vals, idx_hemoglobina)),
            to_text(at(row_vals, idx_gruporiesgo_desc)),
            to_text(at(row_vals, idx_condicion_gestante)),
            to_text(at(row_vals, idx_tipo_edad_pac)),
            to_int(at(row_vals, idx_anio_actual_pac)),
            to_int(at(row_vals, idx_mes_actual_pac)),
            to_int(at(row_vals, idx_dia_actual_pac)),
            to_text(at(row_vals, idx_nombre_establecimiento)),
            to_text(at(row_vals, idx_cie_10)),
            to_text(at(row_vals, idx_diagnostico)),
            to_text(at(row_vals, idx_lab1)),
            to_text(at(row_vals, idx_lab2)),
            to_text(at(row_vals, idx_lab3)),
            to_text(at(row_vals, idx_resultado)),
            to_int(at(row_vals, idx_total)),
        )
        batch.append(rec)

        if len(batch) >= 2000:
            flush()

        now = time.time()
        if now - last_update >= 1.2:
            denom = float(total_rows) if total_rows > 0 else float(max(scanned, 1))
            pct = int(min(99, (scanned / denom) * 100))
            job_update(
                cur,
                job_id,
                progress=pct,
                processed_rows=scanned,
                inserted_rows=inserted,
                message=f"Procesando... {scanned}/{total_rows}",
            )
            db.commit()
            last_update = now

    flush()

    log_line(f"Import done. scanned={scanned}, inserted={inserted}")
    job_update(
        cur,
        job_id,
        status="done",
        progress=100,
        total_rows=scanned,
        processed_rows=scanned,
        inserted_rows=inserted,
        finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        message=f"Importación completada. Insertados: {inserted}",
    )
    db.commit()
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--job", required=True)
    ap.add_argument("--file", required=True)
    args = ap.parse_args()

    job_id = str(args.job).strip()
    file_path = str(args.file).strip()
    if not job_id or not file_path:
        return 2
    if not os.path.exists(file_path):
        return 2
    try:
        return run_import(job_id, file_path)
    except Exception as e:
        try:
            load_dotenv(os.getenv("TAMIZAJE_DOTENV", ".env.local"), override=False)
            load_dotenv(override=False)
            db = connect_db()
            cur = db.cursor()
            ensure_tables(cur)
            db.commit()
            job_update(
                cur,
                job_id,
                status="failed",
                progress=0,
                finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                message="Error: " + (str(e)[:380] if e else "falló"),
            )
            db.commit()
        except Exception:
            pass
        return 1


if __name__ == "__main__":
    sys.exit(main())

