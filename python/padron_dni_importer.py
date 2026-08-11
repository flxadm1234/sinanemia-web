import argparse
import json
import os
import re
import time
from datetime import date, datetime

import mysql.connector
from dotenv import load_dotenv
from openpyxl import load_workbook


def log_line(msg: str):
    p = os.getenv("PADRON_DNI_LOG_PATH", "").strip()
    if not p:
        return
    try:
        with open(p, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.utcnow().isoformat(timespec='seconds')}Z] {msg}\n")
    except Exception:
        pass


def connect_db():
    host = os.getenv("DB_HOST", "127.0.0.1")
    port = int(os.getenv("DB_PORT", "3306"))
    user = os.getenv("DB_USER", "")
    password = os.getenv("DB_PASSWORD", "")
    database = os.getenv("DB_NAME", "")
    return mysql.connector.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        autocommit=False,
    )


def to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def normalize_dni(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        s = str(v)
    elif isinstance(v, float):
        if v != v:
            return None
        s = str(int(v)) if float(int(v)) == float(v) else str(v)
    else:
        s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d]+", "", s)
    return s if s else None


def to_int(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if v != v:
            return None
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d]+", "", s)
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def to_iso_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def normalize_header(v):
    s = str(v or "").strip().upper()
    if not s:
        return ""
    s = s.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").replace("Ñ", "N")
    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s


def find_header_row(ws):
    max_scan = min(40, ws.max_row or 0)
    for i in range(1, max_scan + 1):
        vals = []
        for c in ws[i]:
            t = normalize_header(c.value)
            if t:
                vals.append(t)
        s = set(vals)
        has_ubigeo = any(("UBIGEO" in v) for v in s)
        has_dni = any(("DNI" in v) for v in s) or ("NUMERODEDOCUMENTONACIONALDEIDENTIFICACIONDNI" in s)
        if has_ubigeo and has_dni:
            return i
    return None


def find_col(headers, candidates):
    for idx, h in enumerate(headers):
        if not h:
            continue
        for c in candidates:
            if h == c or h.startswith(c) or c in h:
                return idx
    return None


def best_ubigeo_col(ws, header_row, ubigeo_cols):
    if not ubigeo_cols:
        return None
    max_row = ws.max_row or 0
    data_start = header_row + 1
    scan_end = min(max_row, data_start + 250)
    best = None
    best_score = -1
    for idx in ubigeo_cols:
        score = 0
        for r in range(data_start, scan_end + 1):
            v = ws.cell(row=r, column=idx + 1).value
            u = to_int(v)
            if u and u > 0:
                score += 1
        if score > best_score:
            best_score = score
            best = idx
    return best


def detect_ubigeo_col_t(ws):
    fixed_col_t = 20
    scan_end = 2000
    for r in range(6, scan_end + 1):
        v = ws.cell(row=r, column=fixed_col_t).value
        u = to_int(v)
        if u and u > 0:
            return int(u)
    return None


def pick_sheet(wb):
    for ws in wb.worksheets:
        u = detect_ubigeo_col_t(ws)
        if u:
            return ws
    return wb.worksheets[0]


def sheet_headers(ws, header_row):
    max_col = ws.max_column or 1
    headers = []
    for col in range(1, max_col + 1):
        raw = str(ws.cell(row=header_row, column=col).value or "").strip()
        if raw:
            headers.append(raw)
            continue
        fallback = ""
        for r in range(header_row - 1, 0, -1):
            v = str(ws.cell(row=r, column=col).value or "").strip()
            if v:
                fallback = v
                break
        headers.append(fallback)
    return headers


def period_from_fecha_corte(fecha_corte: date):
    return date(fecha_corte.year, fecha_corte.month, 1)


def job_update(cur, job_id: str, **fields):
    if not fields:
        return
    set_parts = []
    vals = []
    for k, v in fields.items():
        set_parts.append(f"{k}=%s")
        vals.append(v)
    sql = "UPDATE padron_dni_import_jobs SET " + ", ".join(set_parts) + " WHERE id=%s"
    vals.append(job_id)
    cur.execute(sql, vals)


RAW_INSERT_SQL = """
INSERT INTO padron_dni_raw
  (job_id, tipo, row_num, ubigeo, dni, payload)
VALUES
  (%s,%s,%s,%s,%s,%s)
"""


def validate_fecha_corte_rules(cur, job_id: str, ubigeo: int, periodo: date, fecha_corte: date):
    cur.execute(
        """
        SELECT id, fecha_corte
        FROM padron_dni_import_jobs
        WHERE ubigeo = %s
          AND periodo = %s
          AND status IN ('queued','running','done')
          AND id <> %s
        """,
        [ubigeo, periodo, job_id],
    )
    rows = cur.fetchall() or []
    cortes = []
    for r in rows:
        fc = r[1]
        if isinstance(fc, datetime):
            fc = fc.date()
        if isinstance(fc, date):
            cortes.append(fc)
    if len(cortes) >= 2:
        return False, "Ya existen 2 fechas de corte registradas para este ubigeo y mes. Debes eliminar una para volver a cargar."

    inicio = periodo
    if fecha_corte == inicio:
        for fc in cortes:
            if fc == inicio:
                return False, "Ya existe una carga para INICIO DE MES (día 01) en este ubigeo y mes. Debes eliminarla para volver a cargar."
        return True, ""

    for fc in cortes:
        if fc != inicio:
            return False, "Ya existe una carga para AVANCE (segunda fecha de corte) en este ubigeo y mes. Debes eliminarla para volver a cargar."

    return True, ""


def extract_rows(ws, tipo: str):
    header_row = find_header_row(ws)
    if not header_row:
        header_row = 5

    headers = sheet_headers(ws, header_row)
    headers_norm = [normalize_header(h) for h in headers]
    ubigeo_cols = [i for i, h in enumerate(headers_norm) if h and ("UBIGEO" in h)]
    col_ubigeo = best_ubigeo_col(ws, header_row, ubigeo_cols)
    fixed_u = detect_ubigeo_col_t(ws)
    fixed_ubigeo_col = 19
    if (ws.max_column or 0) >= fixed_ubigeo_col + 1:
        v = ws.cell(row=6, column=fixed_ubigeo_col + 1).value
        u = to_int(v)
        if u and u > 0:
            col_ubigeo = fixed_ubigeo_col
    fixed_codpad_col = 2
    fixed_cnv_col = 3
    fixed_dni_col = 5
    col_dni = fixed_dni_col if (ws.max_column or 0) >= fixed_dni_col + 1 else None
    if col_ubigeo is None:
        raise Exception("No se encontró la columna de UBIGEO en la plantilla.")

    ubigeos = set()
    if fixed_u:
        ubigeos.add(int(fixed_u))
    data_start = 6
    empty_run = 0
    out = []

    max_col = max(ws.max_column or 0, len(headers), 70)
    if ws.max_row and ws.max_row >= data_start:
        max_row = ws.max_row
    else:
        max_row = 20000

    for r, row_vals_t in enumerate(
        ws.iter_rows(min_row=data_start, max_row=max_row, max_col=max_col, values_only=True),
        start=data_start,
    ):
        row_vals = list(row_vals_t)
        if len(row_vals) < max_col:
            row_vals.extend([None] * (max_col - len(row_vals)))

        u_cell = to_int(row_vals[col_ubigeo] if col_ubigeo is not None and col_ubigeo < len(row_vals) else None)

        dni = None
        cnv = normalize_dni(row_vals[fixed_cnv_col] if fixed_cnv_col < len(row_vals) else None)
        dni_num = normalize_dni(row_vals[col_dni] if col_dni is not None and col_dni < len(row_vals) else None)
        codpad = normalize_dni(row_vals[fixed_codpad_col] if fixed_codpad_col < len(row_vals) else None)
        dni = cnv or dni_num or codpad

        has_any = False
        for v in row_vals[: min(max_col, 12)]:
            if v is None:
                continue
            s = str(v).strip()
            if s:
                has_any = True
                break
        if not has_any and not dni and not u_cell:
            empty_run += 1
            if empty_run >= 30:
                break
            continue
        empty_run = 0

        u = u_cell
        if (not u or u <= 0) and fixed_u:
            u = fixed_u
        if u and u > 0:
            ubigeos.add(int(u))

        payload = []
        for v in row_vals:
            d = to_iso_date(v)
            if d:
                payload.append(d)
            else:
                payload.append(v if v is not None else "")

        out.append((r, int(u) if u and u > 0 else None, dni, payload))

    if not ubigeos:
        t6 = ws.cell(row=6, column=20).value
        raise Exception(
            f"No se pudo determinar el ubigeo desde el Excel. T6={to_text(t6)} max_row={ws.max_row} max_col={ws.max_column}"
        )

    if len(ubigeos) > 1:
        ubigeos_str = ",".join([str(x) for x in sorted(list(ubigeos))[:10]])
        raise Exception(f"Se detectaron múltiples ubigeos en el Excel ({ubigeos_str}).")

    ubigeo_final = int(list(ubigeos)[0])
    fixed = []
    for row_num, u, dni, payload in out:
        fixed.append((row_num, ubigeo_final, dni, payload))

    return ubigeo_final, headers, fixed


def run_import(
    job_id: str,
    activo_path: str,
    observado_path: str,
    transito_path: str,
    fecha_corte: date,
    update_padron: bool,
    expected_ubigeo=None,
):
    load_dotenv(os.getenv("PADRON_DNI_DOTENV", ".env.local"), override=False)
    load_dotenv(override=False)

    db = connect_db()
    cur = db.cursor()

    job_update(cur, job_id, status="running", progress=0, started_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"), message="Leyendo Excel...")
    db.commit()

    periodo = period_from_fecha_corte(fecha_corte)

    if not activo_path or not str(activo_path).strip():
        raise Exception("Falta el archivo Activo.")

    wb_a = load_workbook(filename=activo_path, read_only=True, data_only=True)
    ws_a = pick_sheet(wb_a)
    ub_a, headers, rows_a = extract_rows(ws_a, "ACTIVO")

    rows_o = []
    rows_t = []
    headers_o = headers
    headers_t = headers
    ub_o = ub_a
    ub_t = ub_a

    if observado_path and str(observado_path).strip():
        wb_o = load_workbook(filename=observado_path, read_only=True, data_only=True)
        ws_o = pick_sheet(wb_o)
        ub_o, headers_o, rows_o = extract_rows(ws_o, "ACTIVO_OBSERVADO")
        if ub_o != ub_a:
            raise Exception(f"Los archivos no tienen el mismo ubigeo. Activo={ub_a} Observado={ub_o}")

    if transito_path and str(transito_path).strip():
        wb_t = load_workbook(filename=transito_path, read_only=True, data_only=True)
        ws_t = pick_sheet(wb_t)
        ub_t, headers_t, rows_t = extract_rows(ws_t, "TRANSITO")
        if ub_t != ub_a:
            raise Exception(f"Los archivos no tienen el mismo ubigeo. Activo={ub_a} Transito={ub_t}")

    if expected_ubigeo and int(expected_ubigeo) != int(ub_a):
        raise Exception(f"El ubigeo del Excel ({ub_a}) no coincide con el ubigeo del usuario ({expected_ubigeo}).")

    if rows_o and len(headers_o) != len(headers):
        raise Exception("Activo y Activo-Observado no tienen la misma plantilla (número de columnas).")
    if rows_t and len(headers_t) != len(headers):
        raise Exception("Activo y Tránsito no tienen la misma plantilla (número de columnas).")

    ok, msg = validate_fecha_corte_rules(cur, job_id, ub_a, periodo, fecha_corte)
    if not ok:
        raise Exception(msg)

    job_update(
        cur,
        job_id,
        ubigeo=ub_a,
        periodo=periodo,
        fecha_corte=fecha_corte,
        headers_json=json.dumps(headers, ensure_ascii=False),
        total_rows=(len(rows_a) + len(rows_o) + len(rows_t)),
        processed_rows=0,
        inserted_rows=0,
        message="Insertando registros...",
    )
    db.commit()

    batch_size = 400
    inserted_total = 0
    processed_total = 0
    total_rows = len(rows_a) + len(rows_o) + len(rows_t)
    last_tick = time.time()

    def insert_many(tipo: str, rows):
        nonlocal inserted_total, processed_total, last_tick
        for i in range(0, len(rows), batch_size):
            part = rows[i : i + batch_size]
            values = []
            for excel_row, ub, dni, payload in part:
                values.append(
                    (
                        job_id,
                        tipo,
                        int(excel_row),
                        int(ub) if ub else None,
                        dni,
                        json.dumps(payload, ensure_ascii=False),
                    )
                )
            if values:
                cur.executemany(RAW_INSERT_SQL, values)
                db.commit()
                inserted_total += cur.rowcount if cur.rowcount and cur.rowcount > 0 else len(values)
            processed_total += len(part)

            now = time.time()
            if now - last_tick >= 0.6 or processed_total >= total_rows:
                progress = int((processed_total / max(1, total_rows)) * 100)
                job_update(
                    cur,
                    job_id,
                    progress=progress,
                    processed_rows=processed_total,
                    inserted_rows=inserted_total,
                    message=f"Procesando... {processed_total}/{total_rows}",
                )
                db.commit()
                last_tick = now

    insert_many("ACTIVO", rows_a)
    if rows_o:
        insert_many("ACTIVO_OBSERVADO", rows_o)
    if rows_t:
        insert_many("TRANSITO", rows_t)

    updated_pn = 0
    if update_padron:
        if fecha_corte != periodo:
            raise Exception("La actualización de padrón nominal solo está permitida para inicio de mes (día 01).")

        job_update(
            cur,
            job_id,
            progress=98,
            processed_rows=processed_total,
            inserted_rows=inserted_total,
            message="Actualizando padrón nominal...",
        )
        db.commit()

    def _s(v):
        if v is None:
            return None
        t = str(v).strip()
        if not t or t.upper() == "NULL":
            return None
        return t

    def _full_name(a, b, c):
        parts = [_s(a), _s(b), _s(c)]
        parts = [p for p in parts if p]
        return " ".join(parts) if parts else None

    def _key_from_payload(payload):
        cnv = normalize_dni(payload[3] if len(payload) > 3 else None)
        dni_num = normalize_dni(payload[5] if len(payload) > 5 else None)
        return cnv or dni_num

    updates_by_key = {}
    for excel_row, ub, dni, payload in rows_a + rows_o + rows_t:
        key = _key_from_payload(payload)
        if not key:
            continue
        updates_by_key[key] = payload

    update_stmt = """
      UPDATE padronnominal pn
      SET
        pn.nombres = CASE
          WHEN pn.nombres IS NULL OR TRIM(pn.nombres) = '' OR UPPER(TRIM(pn.nombres)) = 'NULL' THEN %s
          ELSE pn.nombres
        END,
        pn.appatmadre = CASE
          WHEN pn.appatmadre IS NULL OR TRIM(pn.appatmadre) = '' OR UPPER(TRIM(pn.appatmadre)) = 'NULL' THEN %s
          ELSE pn.appatmadre
        END,
        pn.apmatmadre = CASE
          WHEN pn.apmatmadre IS NULL OR TRIM(pn.apmatmadre) = '' OR UPPER(TRIM(pn.apmatmadre)) = 'NULL' THEN %s
          ELSE pn.apmatmadre
        END,
        pn.nombresmadre = CASE
          WHEN pn.nombresmadre IS NULL OR TRIM(pn.nombresmadre) = '' OR UPPER(TRIM(pn.nombresmadre)) = 'NULL' THEN %s
          ELSE pn.nombresmadre
        END,
        pn.dni_padre = CASE
          WHEN pn.dni_padre IS NULL OR TRIM(pn.dni_padre) = '' OR UPPER(TRIM(pn.dni_padre)) = 'NULL' THEN %s
          ELSE pn.dni_padre
        END,
        pn.nombre_padre = CASE
          WHEN pn.nombre_padre IS NULL OR TRIM(pn.nombre_padre) = '' OR UPPER(TRIM(pn.nombre_padre)) = 'NULL' THEN %s
          ELSE pn.nombre_padre
        END,
        pn.telefonopn = CASE
          WHEN pn.telefonopn IS NULL OR TRIM(pn.telefonopn) = '' OR UPPER(TRIM(pn.telefonopn)) = 'NULL' THEN %s
          ELSE pn.telefonopn
        END
      WHERE
        pn.ubigeo = %s
        AND DATE(pn.etapa) = %s
        AND TRIM(pn.dni) = %s
        AND (
          pn.nombres IS NULL OR TRIM(pn.nombres) = '' OR UPPER(TRIM(pn.nombres)) = 'NULL'
        )
    """

    if update_padron:
        keys = list(updates_by_key.keys())
        if keys:
            batch_u = 400
            done_u = 0
            for i in range(0, len(keys), batch_u):
                part_keys = keys[i : i + batch_u]
                values = []
                for k in part_keys:
                    p = updates_by_key[k]
                    nino_nombre = _full_name(
                        p[8] if len(p) > 8 else None, p[9] if len(p) > 9 else None, p[10] if len(p) > 10 else None
                    )
                    madre_appat = _s(p[45] if len(p) > 45 else None)
                    madre_apmat = _s(p[46] if len(p) > 46 else None)
                    madre_nombres = _s(p[47] if len(p) > 47 else None)
                    madre_cel = normalize_dni(p[48] if len(p) > 48 else None)
                    padre_dni = normalize_dni(p[54] if len(p) > 54 else None)
                    padre_nombre = _full_name(
                        p[55] if len(p) > 55 else None, p[56] if len(p) > 56 else None, p[57] if len(p) > 57 else None
                    )
                    values.append(
                        (
                            nino_nombre,
                            madre_appat,
                            madre_apmat,
                            madre_nombres,
                            padre_dni,
                            padre_nombre,
                            madre_cel,
                            ub_a,
                            periodo,
                            k,
                        )
                    )
                if values:
                    cur.executemany(update_stmt, values)
                    db.commit()
                    updated_pn += cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0

                done_u += len(part_keys)
                progress = 98 + int((done_u / max(1, len(keys))) * 2)
                job_update(
                    cur,
                    job_id,
                    progress=min(progress, 99),
                    message=f"Actualizando padrón nominal... {done_u}/{len(keys)}",
                )
                db.commit()

        log_line(f"[padronnominal] updated_rows={updated_pn} job={job_id} keys={len(keys)}")

    job_update(
        cur,
        job_id,
        status="done",
        progress=100,
        processed_rows=processed_total,
        inserted_rows=inserted_total,
        finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        message=(
            f"Completado. Insertados: {inserted_total}. Padrón actualizado: {updated_pn}"
            if update_padron
            else f"Completado. Insertados: {inserted_total}"
        ),
    )
    db.commit()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--job", required=True)
    ap.add_argument("--activo", required=True)
    ap.add_argument("--observado", default="")
    ap.add_argument("--transito", default="")
    ap.add_argument("--fecha_corte", required=True)
    ap.add_argument("--update_padron", default="0")
    ap.add_argument("--expected_ubigeo", default="")
    args = ap.parse_args()

    try:
        fecha = datetime.strptime(args.fecha_corte.strip(), "%Y-%m-%d").date()
    except Exception:
        raise Exception("Fecha de corte inválida (YYYY-MM-DD).")

    try:
        update_padron = str(args.update_padron or "0").strip() == "1"
        expected_ubigeo = None
        try:
            expected_ubigeo = int(str(args.expected_ubigeo or "").strip()) if str(args.expected_ubigeo or "").strip() else None
        except Exception:
            expected_ubigeo = None
        run_import(args.job, args.activo, args.observado, args.transito, fecha, update_padron, expected_ubigeo)
    except Exception as e:
        try:
            load_dotenv(os.getenv("PADRON_DNI_DOTENV", ".env.local"), override=False)
            load_dotenv(override=False)
            db = connect_db()
            cur = db.cursor()
            job_update(
                cur,
                args.job,
                status="failed",
                progress=0,
                finished_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                message=str(e),
            )
            db.commit()
        except Exception:
            pass
        log_line(f"FAILED: {e}")
        raise


if __name__ == "__main__":
    main()
